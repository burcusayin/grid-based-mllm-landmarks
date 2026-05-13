"""
Dental MLLM Benchmark Pipeline

Reads the benchmark Excel, generates prompts for each query × model × strategy,
submits batch API requests, and collects responses.

Usage:
    # Step 1: Prepare query index (no API key needed)
    python3 pipeline.py prepare

    # Step 2: Submit batches (requires API keys, encodes images on the fly)
    python3 pipeline.py submit

    # Step 3: Check batch status
    python3 pipeline.py status

    # Step 4: Download results
    python3 pipeline.py download

    # Step 5: Parse responses into unified results
    python3 pipeline.py parse
"""

import argparse
import base64
import hashlib
import json
import os
import sys
import tempfile
import time
from pathlib import Path


# ============================================================
# Atomic-write helpers
# ============================================================
# Every persistent write (tracking, downloaded chunks, parsed_responses,
# compliance_stats, query_index) goes through these. The pattern is:
# write to a temp file in the same directory, fsync to disk, then atomically
# rename to the final path. POSIX rename is atomic, so a reader either sees
# the old file or the new file — never a half-written one. If the process
# crashes mid-write, the temp file is left behind but the destination is
# untouched. Resume logic re-runs the write cleanly.

def atomic_write_bytes(path, data):
    """Atomically write bytes to path."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_fd, tmp_name = tempfile.mkstemp(prefix="." + path.name + ".",
                                          dir=str(path.parent))
    try:
        with os.fdopen(tmp_fd, "wb") as f:
            f.write(data)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_name, str(path))
        return path
    except Exception:
        if os.path.exists(tmp_name):
            try:
                os.unlink(tmp_name)
            except OSError:
                pass
        raise


def atomic_write_text(path, text):
    """Atomically write text to path (UTF-8)."""
    return atomic_write_bytes(path, text.encode("utf-8"))


def atomic_write_json(path, obj, **dump_kwargs):
    """Atomically write a JSON-serialisable object to path."""
    dump_kwargs.setdefault("indent", 2)
    return atomic_write_text(path, json.dumps(obj, **dump_kwargs))

import config


# ============================================================
# Excel Parser
# ============================================================

def parse_excel():
    """
    Parse the benchmark Excel file into a list of query dicts.

    Extracted columns (per sheet):
      col  1: Image_ID (derived from row position)
      col  2: Category
      col  3: Target_Structure
      col  4: Question_Prompt (TR)
      col  5: OMFR_1            (first specialist ground truth)
      col  6: OMFR_2            (second specialist ground truth)
      col  7: CONSENSUS_Ground_Truth
      col 10–49: Student_1 .. Student_40

    Returns list of query dicts with keys:
        query_id, sheet, image_id, image_path, category, structure,
        landmark_type, landmark_description_en, uses_fdi, original_prompt_tr,
        omfr_1, omfr_2, consensus_gt, students (list of 40 strings-or-None)
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(config.EXCEL_PATH))
    queries = []

    for sheet_name in ["PANORAMIC", "PERIAPICAL", "CEPHALOMETRIC"]:
        ws = wb[sheet_name]
        grid = config.GRID_SPECS[sheet_name]
        landmarks = config.LANDMARKS[sheet_name]
        lm_per_image = grid["landmarks_per_image"]

        for row_idx in range(2, ws.max_row + 1):
            data_row_num = row_idx - 1
            image_num = -(-data_row_num // lm_per_image)  # ceiling division
            image_id = f"{grid['image_prefix']}_{image_num:03d}"

            category = ws.cell(row_idx, 2).value
            structure = ws.cell(row_idx, 3).value
            prompt_tr = ws.cell(row_idx, 4).value
            omfr_1 = ws.cell(row_idx, 5).value
            omfr_2 = ws.cell(row_idx, 6).value
            consensus_gt = ws.cell(row_idx, 7).value
            students = [ws.cell(row_idx, c).value for c in range(10, 50)]

            if not structure:
                continue

            landmark_def = next(
                (lm for lm in landmarks if lm["structure"] == structure), None
            )
            if not landmark_def:
                print(f"WARNING: Unknown structure '{structure}' in {sheet_name} row {row_idx}")
                continue

            image_path = grid["image_dir"] / f"{image_id}.png"

            queries.append({
                "query_id": f"{image_id}_{structure}",
                "sheet": sheet_name,
                "image_id": image_id,
                "image_path": str(image_path),
                "category": category,
                "structure": structure,
                "landmark_type": landmark_def["type"],
                "landmark_description_en": landmark_def["description_en"],
                "uses_fdi": landmark_def.get("uses_fdi", False),
                "original_prompt_tr": prompt_tr,
                "omfr_1": omfr_1,
                "omfr_2": omfr_2,
                "consensus_gt": consensus_gt,
                "students": students,
            })

    wb.close()
    print(f"Parsed {len(queries)} queries from Excel "
          f"({sum(1 for q in queries if q['sheet'] == 'PANORAMIC')} PAN, "
          f"{sum(1 for q in queries if q['sheet'] == 'PERIAPICAL')} PA, "
          f"{sum(1 for q in queries if q['sheet'] == 'CEPHALOMETRIC')} CEPH)")
    return queries


# ============================================================
# Prompt Generation
# ============================================================

def generate_prompt(query, strategy):
    """Generate (system_prompt, user_prompt) for a query + strategy.

    Strategies:
      zero_shot              — minimal task instruction
      guided                 — guided system prompt + FDI numbering announcement
                               + tooth-number-and-name on FDI-flagged landmarks
      guided_no_tooth_num    — strict-literal of the colleague's prompt-revision
                               proposal (REJECTED by data): identical to `guided`
                               everywhere except on FDI-flagged landmarks, where
                               the user prompt's "tooth #N (anatomic name)" phrase
                               is replaced by just "the anatomic name". Used by
                               the FDI tooth-number ablation in
                               results_ablation_no_tooth_num/.
      guided_patient_left    — patient-frame disambiguation variant (REJECTED by
                               data): identical to `guided` everywhere except on
                               FDI-flagged landmarks, where the parenthetical
                               inside the user prompt (e.g. "(lower left canine)")
                               is rewritten to lead with "patient's left" — the
                               FDI prefix and tooth number are KEPT, only the
                               parenthetical anatomic name is changed. For
                               non-FDI landmarks, byte-identical to `guided`.
                               Used by the Variant A ablation in
                               results_ablation_patient_left/.
      guided_no_LR           — system-prompt diagnostic variant: identical to
                               `guided` everywhere except that the panoramic
                               L–R inversion clause is REMOVED from the system
                               prompt. The user prompt and the rest of the
                               system prompt are byte-identical to `guided`.
                               This is a DIAGNOSTIC, not a candidate canonical
                               prompt (it would also remove the
                               Mental_Foramen_L improvement). For non-PANORAMIC
                               landmarks, byte-identical to `guided` (which is
                               already L–R-clause-free for PA and CEPH). Used
                               by the Variant C ablation in
                               results_ablation_no_LR/.
    """
    sheet = query["sheet"]
    grid = config.GRID_SPECS[sheet]
    lm_type = query["landmark_type"]
    desc = query["landmark_description_en"]
    uses_fdi = query.get("uses_fdi", False)

    # FDI clause for landmarks that reference tooth numbers
    fdi_prefix = "Using the FDI two-digit numbering system, " if uses_fdi else ""
    identify = "identify" if uses_fdi else "Identify"

    if strategy == "zero_shot":
        system_prompt = config.SYSTEM_PROMPT
        if lm_type == "point":
            user_prompt = config.ZERO_SHOT_POINT_TEMPLATE.format(
                cols=grid["cols"], rows=grid["rows"],
                landmark_description=desc,
                fdi_prefix=fdi_prefix, Identify=identify,
            )
        else:
            user_prompt = config.ZERO_SHOT_AREA_TEMPLATE.format(
                cols=grid["cols"], rows=grid["rows"],
                landmark_description=desc,
                fdi_prefix=fdi_prefix, Identify=identify,
            )

    elif strategy in ("guided", "guided_no_tooth_num", "guided_patient_left",
                       "guided_no_LR"):
        # guided_no_LR: identical to guided except the panoramic L–R inversion
        # clause is removed from the system prompt. This is a Variant C
        # diagnostic — it tests whether the L–R clause itself is the proximate
        # cause of the Tooth_33_Apex regression. For non-PANORAMIC modalities
        # the modality clause is already empty, so guided_no_LR produces a
        # system prompt byte-identical to guided.
        if strategy == "guided_no_LR":
            modality_clause = ""
        else:
            modality_clause = config.GUIDED_MODALITY_CLAUSES.get(sheet, "")
        grid_explanation = config.GUIDED_SYSTEM_ADDITION.format(
            cols=grid["cols"],
            max_row=grid["max_row_letter"],
            modality_clause=modality_clause,
        )
        system_prompt = config.SYSTEM_PROMPT + "\n\n" + grid_explanation

        # Description-only overrides for FDI-flagged landmarks. The system
        # prompt and the user-prompt skeleton are identical to `guided`; only
        # the {landmark_description} slot is swapped. For non-FDI landmarks
        # we fall through to the else clause and produce a prompt that is
        # byte-identical to `guided`.
        #
        # Description lookup precedence: query field (newer query_indexes
        # carry it directly) → config.LANDMARKS lookup (older query_indexes
        # don't).
        override_desc = None
        if strategy == "guided_no_tooth_num" and uses_fdi:
            override_desc = query.get("landmark_description_en_no_fdi")
            if not override_desc:
                landmark_def = next(
                    (lm for lm in config.LANDMARKS.get(sheet, [])
                     if lm["structure"] == query.get("structure")),
                    None,
                )
                override_desc = (landmark_def or {}).get("description_en_no_fdi")
            if not override_desc:
                raise ValueError(
                    f"guided_no_tooth_num requires description_en_no_fdi for "
                    f"FDI-flagged landmark {query.get('structure')}; "
                    f"add it to config.LANDMARKS."
                )
        elif strategy == "guided_patient_left" and uses_fdi:
            override_desc = query.get("landmark_description_en_patient_frame")
            if not override_desc:
                landmark_def = next(
                    (lm for lm in config.LANDMARKS.get(sheet, [])
                     if lm["structure"] == query.get("structure")),
                    None,
                )
                override_desc = (landmark_def or {}).get("description_en_patient_frame")
            if not override_desc:
                raise ValueError(
                    f"guided_patient_left requires description_en_patient_frame "
                    f"for FDI-flagged landmark {query.get('structure')}; "
                    f"add it to config.LANDMARKS."
                )

        if override_desc is not None:
            # For guided_no_tooth_num we drop the FDI announcement and use
            # the no-FDI description. For guided_patient_left we KEEP the
            # FDI announcement (and the tooth number inside the parenthetical
            # description that is retained in description_en_patient_frame).
            if strategy == "guided_no_tooth_num":
                effective_desc = override_desc
                effective_fdi_prefix = fdi_prefix   # KEEP per strict-literal proposal
                effective_identify = identify
            elif strategy == "guided_patient_left":
                effective_desc = override_desc
                effective_fdi_prefix = fdi_prefix   # KEEP (patient-frame change is description-only)
                effective_identify = identify
            else:
                effective_desc = override_desc
                effective_fdi_prefix = fdi_prefix
                effective_identify = identify
        else:
            effective_desc = desc
            effective_fdi_prefix = fdi_prefix
            effective_identify = identify

        if lm_type == "point":
            user_prompt = config.GUIDED_POINT_TEMPLATE.format(
                landmark_description=effective_desc,
                fdi_prefix=effective_fdi_prefix, Identify=effective_identify,
            )
        else:
            user_prompt = config.GUIDED_AREA_TEMPLATE.format(
                landmark_description=effective_desc,
                fdi_prefix=effective_fdi_prefix, Identify=effective_identify,
            )
    else:
        raise ValueError(f"Unknown strategy: {strategy}")

    return system_prompt, user_prompt


def encode_image_base64(image_path):
    """Read an image file and return its base64 encoding."""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


# ============================================================
# Image Cache (avoid re-encoding same image for multiple landmarks)
# ============================================================

class ImageCache:
    """Cache base64-encoded images to avoid re-reading from disk."""
    def __init__(self):
        self._cache = {}

    def get(self, image_path):
        if image_path not in self._cache:
            self._cache[image_path] = encode_image_base64(image_path)
        return self._cache[image_path]

    def clear(self):
        self._cache.clear()


# ============================================================
# Request Builders (per provider)
# ============================================================

def build_openai_request(query, strategy, model_cfg, img_b64):
    """Build a single OpenAI batch API request line (dict)."""
    system_prompt, user_prompt = generate_prompt(query, strategy)
    body = {
        "model": model_cfg["model_id"],
        "temperature": model_cfg["temperature"],
        "max_completion_tokens": model_cfg["max_output_tokens"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{img_b64}",
                            "detail": "high",
                        },
                    },
                    {"type": "text", "text": user_prompt},
                ],
            },
        ],
    }
    # Pass seed when configured — best-effort reproducibility. The response
    # will include system_fingerprint so we can detect backend changes.
    if "seed" in model_cfg and model_cfg["seed"] is not None:
        body["seed"] = model_cfg["seed"]
    return {
        "custom_id": f"{query['query_id']}_{strategy}",
        "method": "POST",
        "url": "/v1/chat/completions",
        "body": body,
    }


def build_google_request(query, strategy, model_cfg, img_b64):
    """Build a single Google Gemini batch request (dict)."""
    system_prompt, user_prompt = generate_prompt(query, strategy)
    generation_config = {
        "temperature": model_cfg["temperature"],
        "maxOutputTokens": model_cfg["max_output_tokens"],
        "mediaResolution": model_cfg.get(
            "media_resolution", "MEDIA_RESOLUTION_HIGH"
        ),
    }
    # Gemini accepts a seed integer on generationConfig for reproducibility.
    if "seed" in model_cfg and model_cfg["seed"] is not None:
        generation_config["seed"] = model_cfg["seed"]
    return {
        "custom_id": f"{query['query_id']}_{strategy}",
        "request": {
            "model": model_cfg["model_id"],
            "systemInstruction": {"parts": [{"text": system_prompt}]},
            "contents": [{
                "role": "user",
                "parts": [
                    {"inlineData": {"mimeType": "image/png", "data": img_b64}},
                    {"text": user_prompt},
                ],
            }],
            "generationConfig": generation_config,
        },
    }


def build_anthropic_request(query, strategy, model_cfg, img_b64):
    """Build a single Anthropic Message Batches request (dict)."""
    system_prompt, user_prompt = generate_prompt(query, strategy)
    return {
        "custom_id": f"{query['query_id']}_{strategy}",
        "params": {
            "model": model_cfg["model_id"],
            "max_tokens": model_cfg["max_output_tokens"],
            "temperature": model_cfg["temperature"],
            "system": system_prompt,
            "messages": [{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": img_b64,
                        },
                    },
                    {"type": "text", "text": user_prompt},
                ],
            }],
        },
    }


REQUEST_BUILDERS = {
    "openai": build_openai_request,
    "google": build_google_request,
    "anthropic": build_anthropic_request,
}


# ============================================================
# Prepare Command
# ============================================================

def _select_subset_images(queries, subset_images=None, n_per_modality=None):
    """
    Filter the full query list down to a subset specified by image_id.

    * subset_images: explicit set of image_id strings to keep. Wins over n_per_modality.
    * n_per_modality: deterministic stride selection of N images per modality.
      Uses np.linspace-style indexing to cover the full modality range (first
      and last images always included). This is more informative than taking
      the first N because annotators may have drifted over the course of the
      dataset.

    Returns the filtered queries list and a dict describing what was selected.
    """
    if subset_images is None and n_per_modality is None:
        return queries, None

    # Group queries by (sheet, image_id) preserving order
    from collections import OrderedDict
    by_sheet = OrderedDict()
    for q in queries:
        by_sheet.setdefault(q["sheet"], OrderedDict())
        by_sheet[q["sheet"]].setdefault(q["image_id"], []).append(q)

    keep_ids = set()
    plan = {}

    if subset_images:
        keep_ids = {s.strip() for s in subset_images if s.strip()}
        plan["mode"] = "explicit_subset"
        plan["image_ids"] = sorted(keep_ids)
    else:
        plan["mode"] = "n_per_modality"
        plan["n_per_modality"] = n_per_modality
        plan["per_modality_selection"] = {}
        for sheet, images in by_sheet.items():
            img_list = list(images.keys())  # already in insertion order
            n_total = len(img_list)
            if n_per_modality >= n_total:
                chosen = img_list[:]
            else:
                # Evenly-spaced selection, inclusive of first and last image
                step = (n_total - 1) / (n_per_modality - 1) if n_per_modality > 1 else 0
                indices = sorted({round(i * step) for i in range(n_per_modality)})
                # Pad if rounding collisions reduced the count
                while len(indices) < n_per_modality and len(indices) < n_total:
                    for j in range(n_total):
                        if j not in indices:
                            indices.append(j)
                            break
                indices = sorted(indices)[:n_per_modality]
                chosen = [img_list[i] for i in indices]
            plan["per_modality_selection"][sheet] = chosen
            keep_ids.update(chosen)

    filtered = [q for q in queries if q["image_id"] in keep_ids]
    plan["total_queries_kept"] = len(filtered)
    plan["total_images_kept"] = len(keep_ids)
    return filtered, plan


def cmd_prepare(args):
    """Parse Excel, validate images, and save query index."""
    queries = parse_excel()

    # Optional subset filter (for pilot runs / preliminary experiment).
    # Subset is applied BEFORE image validation so we only require the
    # selected subset to be present on disk.
    subset_images = None
    if getattr(args, "subset_images", None):
        subset_images = [s.strip() for s in args.subset_images.split(",") if s.strip()]
    n_per_modality = getattr(args, "n_per_modality", None)

    queries, plan = _select_subset_images(queries, subset_images, n_per_modality)
    if plan is not None:
        print(f"\nSubset filter applied ({plan['mode']}):")
        if plan["mode"] == "explicit_subset":
            print(f"  image_ids: {plan['image_ids']}")
        else:
            for sheet, imgs in plan["per_modality_selection"].items():
                print(f"  {sheet}: {len(imgs)} images → {imgs}")
        print(f"  Total kept: {plan['total_queries_kept']} queries from "
              f"{plan['total_images_kept']} images")

    # Validate that all SELECTED images exist
    missing = [q for q in queries if not Path(q["image_path"]).exists()]
    if missing:
        print(f"ERROR: {len(missing)} images not found:")
        for q in missing[:5]:
            print(f"  {q['image_path']}")
        sys.exit(1)
    print(f"All {len(queries)} images verified.")

    # Save query index
    config.RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    index_path = config.RESULTS_DIR / "query_index.json"
    atomic_write_json(index_path, queries)
    print(f"Query index saved: {index_path}")

    # Show summary of what will be submitted
    active_models = getattr(args, "models", None) or config.ACTIVE_MODELS
    print(f"\nActive models: {active_models}")
    print(f"Strategies: {config.STRATEGIES}")
    print(f"Total API calls: {len(queries)} × {len(config.ACTIVE_MODELS)} × {len(config.STRATEGIES)} "
          f"= {len(queries) * len(config.ACTIVE_MODELS) * len(config.STRATEGIES)}")


# ============================================================
# Prepare V2 — Final dataset (consensus GT + students + 2 raters w/ washout)
# ============================================================
#
# The v2 prepare path reads the FINAL benchmark Excel
# (data/Final_Dental_MLLM_Benchmark_Data.xlsx) which differs from the original
# in three ways:
#   1. Two specialists rated every query (OMFR_1 + OMFR_2) instead of just one,
#      with intra-rater washout re-ratings on a 180-query subset.
#   2. A team-adjudicated CONSENSUS_Ground_Truth column supersedes OMFR_1 as
#      the canonical GT for downstream analyses.
#   3. The 40 per-student columns are replaced by a single Student_Response
#      column holding the team-adjudicated student consensus per query.
#
# parse_excel_v2 uses HEADER-BASED column lookup (not fixed indices) so the
# parser tolerates further column reorderings without code changes. It is a
# strict superset of parse_excel: every field the legacy schema produced is
# still emitted, with new fields (omfr_1_second, omfr_2_second, student)
# added. The legacy parse_excel is left untouched so results_full/ remains
# byte-reproducible from its original source.
#
# cmd_prepare_v2 writes to config.RESULTS_DIR (typically results_consensus/)
# and additionally cryptographically anchors to results_full/ by recording
# the SHA-256 of every raw OpenAI batch JSONL that the consensus-GT
# re-evaluation will read. Any drift in those frozen outputs is caught at the
# next preflight rather than silently re-scoring a different model snapshot.

V2_REQUIRED_COLUMNS = [
    "Image_ID", "Category", "Target_Structure", "Question_Prompt",
    "OMFR_1", "OMFR_1_Second", "OMFR_2", "OMFR_2_Second",
    "CONSENSUS_Ground_Truth", "Student_Response",
]


def _normalize_cell_value(value):
    """Normalize an Excel cell value to a clean string, or None if blank.
    Whitespace is stripped; values that read as empty after stripping
    become None so downstream code can rely on truthiness."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_excel_v2(excel_path):
    """Parse the FINAL benchmark Excel using header-based column lookup.

    Required columns (any order): Image_ID, Category, Target_Structure,
    Question_Prompt, OMFR_1, OMFR_1_Second, OMFR_2, OMFR_2_Second,
    CONSENSUS_Ground_Truth, Student_Response.

    Returns the same query-dict shape as parse_excel, plus:
        omfr_1_second  — OMFR_1's washout re-rating (None on un-rerated rows)
        omfr_2_second  — OMFR_2's washout re-rating (None on un-rerated rows)
        student        — single team-adjudicated student consensus cell
    The canonical ground-truth field for v2 callers is consensus_gt; omfr_1
    is preserved unchanged so v1 sensitivity comparisons remain possible.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(excel_path))
    queries = []

    for sheet_name in ["PANORAMIC", "PERIAPICAL", "CEPHALOMETRIC"]:
        ws = wb[sheet_name]
        grid = config.GRID_SPECS[sheet_name]
        landmarks = config.LANDMARKS[sheet_name]
        lm_per_image = grid["landmarks_per_image"]

        # Build header map from the first row, validating required columns.
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_index = {h: i + 1 for i, h in enumerate(header) if h is not None}
        missing = [c for c in V2_REQUIRED_COLUMNS if c not in col_index]
        if missing:
            raise ValueError(
                f"{excel_path.name} sheet {sheet_name!r} is missing required "
                f"columns: {missing}. Found columns: {sorted(col_index)}"
            )

        for row_idx in range(2, ws.max_row + 1):
            data_row_num = row_idx - 1
            image_num = -(-data_row_num // lm_per_image)  # ceiling division
            image_id = f"{grid['image_prefix']}_{image_num:03d}"

            structure = _normalize_cell_value(ws.cell(row_idx, col_index["Target_Structure"]).value)
            if not structure:
                continue

            landmark_def = next(
                (lm for lm in landmarks if lm["structure"] == structure), None
            )
            if not landmark_def:
                print(f"WARNING: Unknown structure '{structure}' in {sheet_name} row {row_idx}")
                continue

            image_path = grid["image_dir"] / f"{image_id}.png"

            queries.append({
                "query_id": f"{image_id}_{structure}",
                "sheet": sheet_name,
                "image_id": image_id,
                "image_path": str(image_path),
                "category": _normalize_cell_value(ws.cell(row_idx, col_index["Category"]).value),
                "structure": structure,
                "landmark_type": landmark_def["type"],
                "landmark_description_en": landmark_def["description_en"],
                "landmark_description_en_no_fdi": landmark_def.get("description_en_no_fdi"),
                "landmark_description_en_patient_frame": landmark_def.get("description_en_patient_frame"),
                "uses_fdi": landmark_def.get("uses_fdi", False),
                "original_prompt_tr": _normalize_cell_value(
                    ws.cell(row_idx, col_index["Question_Prompt"]).value),
                "omfr_1":        _normalize_cell_value(ws.cell(row_idx, col_index["OMFR_1"]).value),
                "omfr_1_second": _normalize_cell_value(ws.cell(row_idx, col_index["OMFR_1_Second"]).value),
                "omfr_2":        _normalize_cell_value(ws.cell(row_idx, col_index["OMFR_2"]).value),
                "omfr_2_second": _normalize_cell_value(ws.cell(row_idx, col_index["OMFR_2_Second"]).value),
                "consensus_gt":  _normalize_cell_value(ws.cell(row_idx, col_index["CONSENSUS_Ground_Truth"]).value),
                "student":       _normalize_cell_value(ws.cell(row_idx, col_index["Student_Response"]).value),
            })

    wb.close()
    print(f"Parsed {len(queries)} queries from {excel_path.name} "
          f"({sum(1 for q in queries if q['sheet'] == 'PANORAMIC')} PAN, "
          f"{sum(1 for q in queries if q['sheet'] == 'PERIAPICAL')} PA, "
          f"{sum(1 for q in queries if q['sheet'] == 'CEPHALOMETRIC')} CEPH)")
    return queries


def _sha256_file(path):
    """Return hex SHA-256 of the file at path."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def cmd_prepare_v2(args):
    """Build the consensus-GT query index and anchor to a frozen GPT-5.4 run.

    Reads the FINAL Excel, validates that every required column is populated
    according to the methodology document, and emits an extended query_index
    to config.RESULTS_DIR (sandbox should be results_consensus/, set via the
    DENTAL_MLLM_RESULTS_DIR env var).

    If --anchor-to is given (path to an existing sandbox such as results_full),
    the SHA-256 of every raw JSONL response file under that sandbox's
    run{N}/responses/ is recorded into reanalysis_anchor.json. Subsequent
    re-evaluation scripts can verify these hashes haven't drifted before
    re-scoring against consensus_gt — making it impossible to silently
    re-analyse a different model snapshot.

    No API calls. No writes outside config.RESULTS_DIR.
    """
    excel_path = Path(args.excel) if getattr(args, "excel", None) else (
        config.DATA_DIR / "Final_Dental_MLLM_Benchmark_Data.xlsx"
    )
    if not excel_path.exists():
        print(f"ERROR: Final Excel not found at {excel_path}", file=sys.stderr)
        sys.exit(1)
    excel_sha = _sha256_file(excel_path)
    print(f"Final Excel: {excel_path}")
    print(f"  SHA-256: {excel_sha}")

    queries = parse_excel_v2(excel_path)

    # Sanity checks specific to the v2 schema
    n_total = len(queries)
    expected_total = 900
    if n_total != expected_total:
        print(f"ERROR: expected {expected_total} queries, got {n_total}", file=sys.stderr)
        sys.exit(1)

    # Required-field completeness — these MUST be populated for every row
    for field in ("omfr_1", "omfr_2", "consensus_gt", "student"):
        nulls = [q["query_id"] for q in queries if not q.get(field)]
        if nulls:
            print(f"ERROR: {len(nulls)} queries have null {field}; first 5: {nulls[:5]}",
                  file=sys.stderr)
            sys.exit(1)
    print(f"Required-field completeness OK (omfr_1, omfr_2, consensus_gt, student all populated).")

    # OMFR_*_Second is partial by design (washout-period subset)
    n_o1s = sum(1 for q in queries if q.get("omfr_1_second"))
    n_o2s = sum(1 for q in queries if q.get("omfr_2_second"))
    print(f"Intra-rater coverage: omfr_1_second populated for {n_o1s}/{n_total} queries; "
          f"omfr_2_second for {n_o2s}/{n_total}.")

    # Image existence
    missing = [q for q in queries if not Path(q["image_path"]).exists()]
    if missing:
        print(f"ERROR: {len(missing)} images not found:", file=sys.stderr)
        for q in missing[:5]:
            print(f"  {q['image_path']}", file=sys.stderr)
        sys.exit(1)
    print(f"All {n_total} images verified.")

    # Save extended query index
    config.RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    index_path = config.RESULTS_DIR / "query_index.json"
    atomic_write_json(index_path, queries)
    print(f"Extended query index saved: {index_path}")

    # Manifest
    manifest = {
        "schema_version": "v2_consensus",
        "source_excel": str(excel_path),
        "source_excel_sha256": excel_sha,
        "n_queries": n_total,
        "n_intra_rater_omfr1": n_o1s,
        "n_intra_rater_omfr2": n_o2s,
        "canonical_gt_field": "consensus_gt",
        "preserved_fields_for_sensitivity": ["omfr_1"],
        "produced_by": "pipeline.cmd_prepare_v2",
    }
    atomic_write_json(config.RESULTS_DIR / "v2_manifest.json", manifest)

    # Cryptographic anchor to a frozen GPT-5.4 run (e.g. results_full/)
    anchor_to = getattr(args, "anchor_to", None)
    if anchor_to:
        anchor_root = Path(anchor_to).resolve()
        if not anchor_root.exists():
            print(f"ERROR: --anchor-to path does not exist: {anchor_root}", file=sys.stderr)
            sys.exit(1)
        if anchor_root == config.RESULTS_DIR.resolve():
            print(f"ERROR: cannot anchor to the v2 sandbox itself "
                  f"({anchor_root}); pick an upstream frozen run.", file=sys.stderr)
            sys.exit(1)
        anchor = {"anchor_root": str(anchor_root), "files": {}}
        n_anchored = 0
        for rep_dir in sorted(anchor_root.glob("run*/responses")):
            for jl in sorted(rep_dir.glob("*.jsonl")):
                rel = jl.relative_to(anchor_root)
                anchor["files"][str(rel)] = {
                    "sha256": _sha256_file(jl),
                    "size": jl.stat().st_size,
                }
                n_anchored += 1
        # Also anchor the upstream query_index, if present, since it defines
        # the (custom_id → image/landmark) mapping the responses encode.
        upstream_qi = anchor_root / "query_index.json"
        if upstream_qi.exists():
            anchor["upstream_query_index"] = {
                "path": str(upstream_qi.relative_to(anchor_root)),
                "sha256": _sha256_file(upstream_qi),
                "size": upstream_qi.stat().st_size,
            }
        atomic_write_json(config.RESULTS_DIR / "reanalysis_anchor.json", anchor)
        print(f"Anchored {n_anchored} JSONL files from {anchor_root} → "
              f"{config.RESULTS_DIR/'reanalysis_anchor.json'}")
    else:
        print("(no --anchor-to specified; skipping JSONL anchor)")

    print(f"\nv2 sandbox ready at {config.RESULTS_DIR}")


# ============================================================
# Prepare Ablation — focused subset for the FDI×L-R interaction test
# ============================================================
#
# This builds a tightly-filtered query_index for the FDI ablation experiment.
# It is explicitly designed so the user CANNOT accidentally re-run the entire
# 900-query benchmark: the prepare step asserts on (a) the requested
# structures filter is non-empty, (b) the resulting query count matches the
# expected number for that filter, and (c) the sandbox is not a v1/v2
# directory. The submit step inherits the standard `.api_lock`, per-chunk
# persistence, atomic writes, and SHA anchoring.
#
# Cryptographic anchoring: the ablation sandbox records SHAs of the Final
# Excel, the v2 query_index, and (if --anchor-to is given) every raw JSONL
# at that path. The downstream analysis script verifies all of these before
# scoring or comparing.

def cmd_prepare_ablation(args):
    """Build a focused query_index for an ablation experiment.

    Filters parse_excel_v2's output by --structures (comma-separated) and
    writes the result to config.RESULTS_DIR/query_index.json plus a manifest
    describing the filter and the ablation provenance.

    Refuses if:
      - the resulting filter is empty
      - the resulting count differs from --expected-count (when supplied)
      - config.RESULTS_DIR resolves to a known v1/v2 sandbox or `data/`
      - the Final Excel is missing or unreadable

    No API calls. No writes outside config.RESULTS_DIR.
    """
    # ── Hard sandbox isolation ─────────────────────────────────────
    sandbox = config.RESULTS_DIR.resolve()
    forbidden = []
    for name in ("results_full", "results_consensus", "data"):
        p = (config.PROJECT_ROOT / name).resolve()
        if sandbox == p or p in sandbox.parents:
            forbidden.append(str(p))
    if forbidden:
        print(f"ERROR: sandbox {sandbox} is or lives inside a frozen/canonical "
              f"directory ({forbidden}). Refusing to prepare ablation here.",
              file=sys.stderr)
        sys.exit(1)

    # ── Parse args ─────────────────────────────────────────────────
    excel_path = Path(args.excel) if getattr(args, "excel", None) else (
        config.DATA_DIR / "Final_Dental_MLLM_Benchmark_Data.xlsx"
    )
    if not excel_path.exists():
        print(f"ERROR: Final Excel not found at {excel_path}", file=sys.stderr)
        sys.exit(1)

    structures_arg = (getattr(args, "structures", None) or "").strip()
    if not structures_arg:
        print(f"ERROR: --structures is required (comma-separated, e.g. "
              f"'Tooth_33_Apex'); refusing to prepare an unfiltered ablation.",
              file=sys.stderr)
        sys.exit(1)
    target_structures = [s.strip() for s in structures_arg.split(",") if s.strip()]

    # ── Load + filter ──────────────────────────────────────────────
    excel_sha = _sha256_file(excel_path)
    print(f"Final Excel: {excel_path}")
    print(f"  SHA-256: {excel_sha}")

    all_queries = parse_excel_v2(excel_path)
    queries = [q for q in all_queries if q["structure"] in target_structures]
    print(f"Filter applied (structures = {target_structures}): "
          f"{len(queries)} queries kept (of {len(all_queries)})")
    if not queries:
        print(f"ERROR: filter produced 0 queries; check --structures spelling.",
              file=sys.stderr)
        sys.exit(1)

    expected_count = getattr(args, "expected_count", None)
    if expected_count is not None and len(queries) != expected_count:
        print(f"ERROR: expected {expected_count} queries after filter, "
              f"got {len(queries)}.", file=sys.stderr)
        sys.exit(1)

    # ── Field completeness ─────────────────────────────────────────
    for field in ("consensus_gt", "omfr_1", "omfr_2", "student"):
        nulls = [q["query_id"] for q in queries if not q.get(field)]
        if nulls:
            print(f"ERROR: {len(nulls)} queries have null {field}; "
                  f"first 5: {nulls[:5]}", file=sys.stderr)
            sys.exit(1)

    # If any FDI landmark, the alternative-description fields must be populated
    # for the ablation strategies to work (guided_no_tooth_num needs
    # description_en_no_fdi; guided_patient_left needs
    # description_en_patient_frame). We require BOTH because the same prepared
    # sandbox may be used to render either variant.
    for q in queries:
        if q.get("uses_fdi"):
            for field in ("landmark_description_en_no_fdi",
                          "landmark_description_en_patient_frame"):
                if not q.get(field):
                    print(f"ERROR: FDI landmark {q['structure']} is missing "
                          f"{field}; add it to config.LANDMARKS.",
                          file=sys.stderr)
                    sys.exit(1)
    print("Required-field completeness OK.")

    # Image existence
    missing = [q for q in queries if not Path(q["image_path"]).exists()]
    if missing:
        print(f"ERROR: {len(missing)} images not found.", file=sys.stderr)
        sys.exit(1)
    print(f"All {len(queries)} images verified.")

    # ── Save outputs ───────────────────────────────────────────────
    config.RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    index_path = config.RESULTS_DIR / "query_index.json"
    atomic_write_json(index_path, queries)
    print(f"Filtered query index saved: {index_path}")

    manifest = {
        "schema_version": "v2_consensus_ablation",
        "ablation_label": getattr(args, "label", "fdi_ablation"),
        "source_excel": str(excel_path),
        "source_excel_sha256": excel_sha,
        "filter_structures": target_structures,
        "n_queries": len(queries),
        "canonical_gt_field": "consensus_gt",
        "produced_by": "pipeline.cmd_prepare_ablation",
    }
    atomic_write_json(config.RESULTS_DIR / "ablation_manifest.json", manifest)

    # Cryptographic anchor to a frozen run (e.g. results_full/) for the
    # zero-shot + guided baseline data we will compare against, AND to the
    # v2 query_index as the canonical source of consensus_gt.
    anchor_to = getattr(args, "anchor_to", None)
    v2_index = getattr(args, "v2_index", None) or (
        config.PROJECT_ROOT / "results_consensus" / "query_index.json"
    )
    anchor = {
        "final_excel_sha256": excel_sha,
    }
    if Path(v2_index).exists():
        anchor["v2_query_index"] = {
            "path": str(v2_index),
            "sha256": _sha256_file(Path(v2_index)),
            "size": Path(v2_index).stat().st_size,
        }
    if anchor_to:
        anchor_root = Path(anchor_to).resolve()
        if not anchor_root.exists():
            print(f"ERROR: --anchor-to path does not exist: {anchor_root}",
                  file=sys.stderr)
            sys.exit(1)
        if anchor_root == config.RESULTS_DIR.resolve():
            print(f"ERROR: cannot anchor to the ablation sandbox itself.",
                  file=sys.stderr)
            sys.exit(1)
        # Filter anchor to JSONLs that contain at least one custom_id matching
        # our target structures (i.e. the chunks we'll need to re-read).
        target_query_ids = {q["query_id"] for q in queries}
        anchor["anchor_root"] = str(anchor_root)
        anchor["files"] = {}
        n_total, n_relevant = 0, 0
        for rep_dir in sorted(anchor_root.glob("run*/responses")):
            for jl in sorted(rep_dir.glob("*.jsonl")):
                n_total += 1
                # peek first line to check if any custom_id matches our filter
                try:
                    with open(jl) as f:
                        first = f.readline()
                    if not first.strip():
                        continue
                    obj = json.loads(first)
                    cid = obj.get("custom_id", "")
                    qid = cid.rsplit("_", 1)[0] if "_" in cid else cid
                    # Cheaper: scan all lines once and short-circuit
                    relevant = False
                    with open(jl) as f:
                        for line in f:
                            if not line.strip():
                                continue
                            try:
                                cid = json.loads(line)["custom_id"]
                            except Exception:
                                continue
                            qid = (cid[:-len("_zero_shot")] if cid.endswith("_zero_shot")
                                   else cid[:-len("_guided")] if cid.endswith("_guided")
                                   else None)
                            if qid in target_query_ids:
                                relevant = True
                                break
                except Exception as e:
                    print(f"WARN: could not scan {jl}: {e}", file=sys.stderr)
                    continue
                if relevant:
                    rel = jl.relative_to(anchor_root)
                    anchor["files"][str(rel)] = {
                        "sha256": _sha256_file(jl),
                        "size": jl.stat().st_size,
                    }
                    n_relevant += 1
        print(f"Anchored {n_relevant} relevant JSONL files (of {n_total} scanned) "
              f"from {anchor_root}")
    atomic_write_json(config.RESULTS_DIR / "ablation_anchor.json", anchor)
    print(f"\nAblation sandbox ready at {config.RESULTS_DIR}")


# ============================================================
# Submit Command
# ============================================================

# OpenAI batch file size limit ~200MB; each request is ~3MB → ~60 per chunk
OPENAI_CHUNK_SIZE = 50
# Google batchGenerateContent: max 100 requests per call
GOOGLE_CHUNK_SIZE = 100
# Anthropic: 256 MB total batch body limit. With ~3 MB base64-encoded PNG
# images per panoramic request, 50 requests/chunk ≈ 150 MB body (well under
# the 256 MB hard cap). 900 panoramic queries → 18 chunks per strategy.
ANTHROPIC_CHUNK_SIZE = 50


_OPENAI_TRANSIENT_HTTP = {401, 408, 429, 500, 502, 503, 504}


def _openai_call_with_retries(fn, *args, _label="openai call", _max_attempts=4, **kwargs):
    """
    Retry wrapper for OpenAI HTTP calls. Retries on transient HTTP errors
    (including 401, which we have observed as a transient OpenAI auth flake)
    with exponential backoff: 2s, 4s, 8s. Re-raises after _max_attempts.
    """
    import urllib.error
    last_err = None
    for attempt in range(1, _max_attempts + 1):
        try:
            return fn(*args, **kwargs)
        except urllib.error.HTTPError as e:
            if e.code not in _OPENAI_TRANSIENT_HTTP or attempt == _max_attempts:
                raise
            last_err = e
            wait = 2 ** attempt
            print(f"    {_label}: HTTP {e.code} on attempt {attempt}/{_max_attempts}; "
                  f"retrying in {wait}s")
            time.sleep(wait)
        except (urllib.error.URLError, ConnectionError) as e:
            if attempt == _max_attempts:
                raise
            last_err = e
            wait = 2 ** attempt
            print(f"    {_label}: {type(e).__name__} on attempt {attempt}/{_max_attempts}; "
                  f"retrying in {wait}s")
            time.sleep(wait)
    raise last_err  # pragma: no cover (loop always returns or raises)


def upload_openai_file(jsonl_path, api_key):
    """Upload a JSONL file to OpenAI and return the file ID."""
    import urllib.request

    boundary = "----BatchBoundary"
    file_data = Path(jsonl_path).read_bytes()
    filename = Path(jsonl_path).name

    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="purpose"\r\n\r\nbatch\r\n'
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        f"Content-Type: application/jsonl\r\n\r\n"
    ).encode() + file_data + f"\r\n--{boundary}--\r\n".encode()

    req = urllib.request.Request(
        "https://api.openai.com/v1/files",
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": f"multipart/form-data; boundary={boundary}",
        },
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read())["id"]


def create_openai_batch(file_id, model_key, api_key):
    """Create an OpenAI batch from an uploaded file."""
    import urllib.request

    body = json.dumps({
        "input_file_id": file_id,
        "endpoint": "/v1/chat/completions",
        "completion_window": "24h",
        "metadata": {"model": model_key},
    }).encode()

    req = urllib.request.Request(
        "https://api.openai.com/v1/batches",
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def submit_openai(queries, model_key, strategy, api_key, *,
                   on_chunk_success=None):
    """
    Submit queries to OpenAI Batch API in chunks.

    on_chunk_success: optional callback(chunk_num, batch_id) invoked after
    each chunk's batch is successfully created at OpenAI. The caller can use
    this to persist partial state to disk between chunks, so that a crash
    mid-bundle does not strand already-created batch_ids without tracking.
    """
    model_cfg = config.MODELS[model_key]
    img_cache = ImageCache()
    batch_ids = []

    for chunk_idx in range(0, len(queries), OPENAI_CHUNK_SIZE):
        chunk = queries[chunk_idx:chunk_idx + OPENAI_CHUNK_SIZE]
        chunk_num = chunk_idx // OPENAI_CHUNK_SIZE + 1
        total_chunks = -(-len(queries) // OPENAI_CHUNK_SIZE)

        print(f"  Chunk {chunk_num}/{total_chunks} ({len(chunk)} requests)...")

        # Write chunk to temp JSONL file
        tmp_path = None
        try:
            # Assign tmp_path IMMEDIATELY so the finally block can clean up
            # if the write loop raises (e.g., missing image file). The prior
            # version assigned tmp_path only after the loop completed, which
            # leaked temp files on mid-loop exceptions.
            with tempfile.NamedTemporaryFile(mode="w", suffix=".jsonl", delete=False) as tmp:
                tmp_path = tmp.name
                for query in chunk:
                    img_b64 = img_cache.get(query["image_path"])
                    req = build_openai_request(query, strategy, model_cfg, img_b64)
                    tmp.write(json.dumps(req) + "\n")

            file_id = _openai_call_with_retries(
                upload_openai_file, tmp_path, api_key,
                _label=f"upload chunk{chunk_num}")
            print(f"    Uploaded file: {file_id}")
            batch_resp = _openai_call_with_retries(
                create_openai_batch, file_id, model_key, api_key,
                _label=f"create batch chunk{chunk_num}")
            batch_id = batch_resp["id"]
            print(f"    Batch created: {batch_id}")
            batch_ids.append(batch_id)

            # Persist partial state immediately so a later-chunk crash does
            # not strand earlier chunks' batch_ids.
            if on_chunk_success is not None:
                try:
                    on_chunk_success(chunk_num, batch_id, list(batch_ids))
                except Exception as cb_err:
                    # Persistence failure should not lose the chunk; warn but
                    # continue. The caller will see the partial state on resume.
                    print(f"    WARN: on_chunk_success callback failed: {cb_err}",
                          file=sys.stderr)
        finally:
            if tmp_path is not None and os.path.exists(tmp_path):
                os.unlink(tmp_path)

        # Clear image cache after every chunk to bound peak memory.
        img_cache.clear()

    return batch_ids


def submit_google(queries, model_key, strategy, api_key):
    """Submit queries to Google Gemini batchGenerateContent API."""
    import urllib.request

    model_cfg = config.MODELS[model_key]
    img_cache = ImageCache()
    chunk_results = []

    config.RESPONSES_DIR.mkdir(parents=True, exist_ok=True)
    batch_name = f"{model_key}_{strategy}"

    for chunk_idx in range(0, len(queries), GOOGLE_CHUNK_SIZE):
        chunk = queries[chunk_idx:chunk_idx + GOOGLE_CHUNK_SIZE]
        chunk_num = chunk_idx // GOOGLE_CHUNK_SIZE + 1
        total_chunks = -(-len(queries) // GOOGLE_CHUNK_SIZE)

        print(f"  Chunk {chunk_num}/{total_chunks} ({len(chunk)} requests)...")

        # Build requests for this chunk
        custom_ids = []
        batch_requests = []
        for query in chunk:
            img_b64 = img_cache.get(query["image_path"])
            req_data = build_google_request(query, strategy, model_cfg, img_b64)
            custom_ids.append(req_data["custom_id"])
            batch_requests.append(req_data["request"])

        body = json.dumps({"requests": batch_requests}).encode()
        model_id = model_cfg["model_id"]
        url = (
            f"https://generativelanguage.googleapis.com/v1beta/"
            f"models/{model_id}:batchGenerateContent?key={api_key}"
        )

        req = urllib.request.Request(
            url, data=body,
            headers={"Content-Type": "application/json"},
        )

        try:
            with urllib.request.urlopen(req, timeout=600) as resp:
                result = json.loads(resp.read())

            # Pair responses with custom IDs
            responses = result.get("responses", [])
            if len(responses) != len(custom_ids):
                print(f"    WARNING: Google returned {len(responses)} responses "
                      f"for {len(custom_ids)} requests in chunk {chunk_num}. "
                      f"Missing requests will be recorded as None.")
            paired = []
            for i, cid in enumerate(custom_ids):
                paired.append({
                    "custom_id": cid,
                    "response": responses[i] if i < len(responses) else None,
                })

            chunk_path = config.RESPONSES_DIR / f"{batch_name}_chunk{chunk_num:03d}.json"
            atomic_write_json(chunk_path, paired)
            chunk_results.append(str(chunk_path))
            print(f"    Saved: {chunk_path}")

        except Exception as e:
            print(f"    ERROR: {e}")
            chunk_results.append(f"ERROR: {e}")

        # Rate limiting between chunks
        if chunk_num < total_chunks:
            time.sleep(2)

    return chunk_results


def submit_anthropic(queries, model_key, strategy, api_key):
    """Submit queries to Anthropic Message Batches API in chunks.

    Anthropic enforces a 256 MB total body size per batch. With base64-encoded
    PNG images averaging ~3 MB per request (raw image × ~1.34 b64 expansion +
    JSON overhead), 900 panoramic requests would be ~2.7 GB — 10.6× over the
    limit. We chunk into ANTHROPIC_CHUNK_SIZE requests per batch, keeping the
    per-batch body well under 256 MB (~150 MB at chunk_size=50).

    Returns a list of batch_ids (one per chunk), in order. Downstream code
    in cmd_status / cmd_download iterates over this list.
    """
    import urllib.request

    model_cfg = config.MODELS[model_key]
    img_cache = ImageCache()
    batch_ids: list[str] = []
    n = len(queries)
    total_chunks = -(-n // ANTHROPIC_CHUNK_SIZE)

    for chunk_idx in range(0, n, ANTHROPIC_CHUNK_SIZE):
        chunk = queries[chunk_idx:chunk_idx + ANTHROPIC_CHUNK_SIZE]
        chunk_num = chunk_idx // ANTHROPIC_CHUNK_SIZE + 1

        requests_list = []
        for q in chunk:
            img_b64 = img_cache.get(q["image_path"])
            requests_list.append(build_anthropic_request(q, strategy, model_cfg, img_b64))
        # Free image cache between chunks (b64 strings are still in requests_list
        # but cache stops growing for the next chunk).
        img_cache.clear()

        body = json.dumps({"requests": requests_list}).encode()
        body_mb = len(body) / 1e6
        print(f"  Chunk {chunk_num}/{total_chunks}: submitting "
              f"{len(requests_list)} requests ({body_mb:.1f} MB body)...")
        if body_mb > 240:
            # Hard refusal before we hit Anthropic's 256 MB limit and waste a call.
            raise RuntimeError(
                f"Anthropic chunk body is {body_mb:.1f} MB — exceeds safe 240 MB "
                f"headroom under the 256 MB hard limit. Lower ANTHROPIC_CHUNK_SIZE."
            )

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages/batches",
            data=body,
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
        )
        with urllib.request.urlopen(req, timeout=300) as resp:
            batch_resp = json.loads(resp.read())
        batch_id = batch_resp["id"]
        batch_ids.append(batch_id)
        print(f"    Batch created: {batch_id}")

        # Small pause between chunks to avoid hammering the API
        if chunk_num < total_chunks:
            time.sleep(2)

    return batch_ids


def cmd_submit(args):
    """Submit batches for all active models × strategies."""
    # Safety gate: .api_lock prevents accidental API spend
    api_lock = Path(__file__).resolve().parent / ".api_lock"
    if api_lock.exists():
        print("ERROR: API calls are locked. Delete .api_lock to unlock.")
        sys.exit(1)

    # Load query index
    index_path = config.RESULTS_DIR / "query_index.json"
    if not index_path.exists():
        print("ERROR: Query index not found. Run 'prepare' first.")
        sys.exit(1)

    with open(index_path) as f:
        queries = json.load(f)

    config.RESPONSES_DIR.mkdir(parents=True, exist_ok=True)

    # Determine which models to submit. --models overrides config.ACTIVE_MODELS
    # for this run only (no file mutation). Every requested model must be
    # defined in config.MODELS; otherwise we refuse to submit so no budget
    # is spent on a typo.
    override_models = None
    if getattr(args, "models", None):
        override_models = [m.strip() for m in args.models.split(",") if m.strip()]
        bad = [m for m in override_models if m not in config.MODELS]
        if bad:
            print(f"ERROR: unknown model(s) in --models: {bad}")
            print(f"Valid models: {sorted(config.MODELS.keys())}")
            sys.exit(1)
    models_to_run = override_models if override_models is not None else config.ACTIVE_MODELS

    # Determine which strategies to submit. --strategies overrides config.STRATEGIES
    # for this run only (no file mutation). Used for ablation experiments where
    # we want to submit only `guided_no_tooth_num` without touching the canonical pair.
    # Every requested strategy must be one we know how to render in
    # generate_prompt; we validate by rendering a test prompt for the first query.
    override_strategies = None
    if getattr(args, "strategies", None):
        override_strategies = [s.strip() for s in args.strategies.split(",") if s.strip()]
        for s in override_strategies:
            try:
                _ = generate_prompt(queries[0], s)
            except Exception as e:
                print(f"ERROR: strategy {s!r} cannot be rendered by generate_prompt: {e}",
                      file=sys.stderr)
                sys.exit(1)
    strategies_to_run = override_strategies if override_strategies is not None else config.STRATEGIES

    # Load/create tracking file
    tracking_path = config.RESULTS_DIR / "batch_tracking.json"
    tracking = {}
    if tracking_path.exists():
        with open(tracking_path) as f:
            tracking = json.load(f)

    print(f"\nSubmitting to models: {models_to_run}")
    print(f"Strategies:          {strategies_to_run}")
    print(f"Total API calls:     {len(queries) * len(models_to_run) * len(strategies_to_run)}")

    for model_key in models_to_run:
        model_cfg = config.MODELS[model_key]
        provider = model_cfg["provider"]

        try:
            api_key = config.get_api_key(provider)
        except ValueError as e:
            print(f"\nSKIP {model_key}: {e}")
            continue

        for strategy in strategies_to_run:
            batch_name = f"{model_key}_{strategy}"

            # Skip if this bundle has already been fully submitted. "partial"
            # status (some chunks created at OpenAI but submission did not
            # complete) is treated like "failed" — we re-submit the whole bundle.
            # The orphaned partial batches at OpenAI will continue processing;
            # operator should manually cancel them if needed.
            existing_status = tracking.get(batch_name, {}).get("status")
            if batch_name in tracking and existing_status not in ("failed", "partial"):
                print(f"\nSKIP: {batch_name} already submitted")
                continue
            if existing_status == "partial":
                stranded = tracking[batch_name].get("batch_ids", [])
                print(f"\nWARN: {batch_name} previously crashed mid-submit; "
                      f"{len(stranded)} stranded batch(es) at OpenAI:")
                for bid in stranded:
                    print(f"  {bid}  (consider manual cancel via API)")
                print(f"  Re-submitting full bundle.")

            print(f"\n{'='*60}")
            print(f"Submitting: {batch_name} ({provider}, {len(queries)} queries)")
            print(f"{'='*60}")

            try:
                if provider == "openai":
                    # Per-chunk persistence: if submit_openai crashes mid-bundle,
                    # we still keep the batch_ids of chunks that successfully
                    # made it to OpenAI. This prevents stranding (and being
                    # billed for) orphaned batches with no local record.
                    def _persist_chunk(chunk_num, batch_id, batch_ids_so_far):
                        tracking[batch_name] = {
                            "provider": provider,
                            "batch_ids": list(batch_ids_so_far),
                            "status": "partial",  # not yet "submitted"
                            "model": model_key,
                            "strategy": strategy,
                            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                        }
                        atomic_write_json(tracking_path, tracking)
                    batch_ids = submit_openai(
                        queries, model_key, strategy, api_key,
                        on_chunk_success=_persist_chunk)
                    tracking[batch_name] = {
                        "provider": provider,
                        "batch_ids": batch_ids,
                        "status": "submitted",
                        "model": model_key,
                        "strategy": strategy,
                        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    }
                elif provider == "google":
                    chunk_paths = submit_google(queries, model_key, strategy, api_key)
                    # Surface any per-chunk errors that submit_google logged
                    # as "ERROR: ..." entries in the chunk_paths list. Without
                    # this, status was unconditionally "completed" even when
                    # individual chunks failed, masking partial-failure
                    # bundles from downstream code.
                    chunk_errors = [p for p in chunk_paths
                                    if isinstance(p, str) and p.startswith("ERROR:")]
                    n_ok = len(chunk_paths) - len(chunk_errors)
                    if not chunk_errors:
                        gstatus = "completed"
                    elif n_ok > 0:
                        gstatus = "completed_with_failures"
                    else:
                        gstatus = "failed"
                    tracking[batch_name] = {
                        "provider": provider,
                        "chunk_paths": chunk_paths,
                        "chunk_errors": chunk_errors,  # for orchestrator inspection
                        "n_chunks_ok": n_ok,
                        "n_chunks_failed": len(chunk_errors),
                        "status": gstatus,
                        "model": model_key,
                        "strategy": strategy,
                        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    }
                elif provider == "anthropic":
                    batch_ids = submit_anthropic(queries, model_key, strategy, api_key)
                    tracking[batch_name] = {
                        "provider": provider,
                        "batch_ids": batch_ids,
                        "status": "submitted",
                        "model": model_key,
                        "strategy": strategy,
                        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    }
            except Exception as e:
                print(f"  FATAL ERROR: {e}")
                tracking[batch_name] = {
                    "provider": provider,
                    "status": "failed",
                    "error": str(e),
                    "model": model_key,
                    "strategy": strategy,
                    "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                }

            # Save tracking after each submission (atomic)
            atomic_write_json(tracking_path, tracking)

    print(f"\nTracking saved: {tracking_path}")


# ============================================================
# Status Command
# ============================================================

def cmd_status(args):
    """Check status of all submitted batches."""
    import urllib.request

    tracking_path = config.RESULTS_DIR / "batch_tracking.json"
    if not tracking_path.exists():
        print("No batches submitted yet. Run 'submit' first.")
        return

    with open(tracking_path) as f:
        tracking = json.load(f)

    print(f"\n{'Batch':<40} {'Provider':<12} {'Status':<15} {'Details'}")
    print("-" * 90)

    for batch_name, info in tracking.items():
        provider = info["provider"]
        status = info.get("status", "unknown")

        if status in ("completed", "failed"):
            details = info.get("error", "")
            print(f"{batch_name:<40} {provider:<12} {status:<15} {details}")
            continue

        # Check OpenAI batch status
        if provider == "openai":
            try:
                api_key = config.get_api_key("openai")
            except Exception as e:
                # Bundle-level error (config / auth) — surface and skip this bundle
                print(f"{batch_name:<40} {provider:<12} {'error':<15} {e}",
                      file=sys.stderr)
                continue

            # Per-chunk progress is preserved across polls so a transient
            # failure on one chunk does not reset the others, and successful
            # chunks are not re-polled unnecessarily.
            chunk_progress = info.setdefault("chunk_progress", {})
            chunk_errors = []
            completed_count = 0
            total_count = 0

            for batch_id in info.get("batch_ids", []):
                if batch_id.startswith("ERROR"):
                    continue

                # If this chunk is already known-completed with an output file,
                # trust the previous record and skip the API call.
                prev = chunk_progress.get(batch_id, {})
                if prev.get("status") == "completed" and prev.get("output_file_id"):
                    completed_count += prev.get("completed_requests", 0)
                    total_count += prev.get("total_requests", 0)
                    continue

                def _poll_openai_batch(_bid=batch_id):
                    req = urllib.request.Request(
                        f"https://api.openai.com/v1/batches/{_bid}",
                        headers={"Authorization": f"Bearer {api_key}"},
                    )
                    with urllib.request.urlopen(req) as resp:
                        return json.loads(resp.read())

                # Per-chunk error handling: log and continue with siblings
                # rather than aborting the whole bundle's status update.
                try:
                    batch_resp = _openai_call_with_retries(
                        _poll_openai_batch,
                        _label=f"poll {batch_name}/{batch_id[:20]}")
                except Exception as e:
                    msg = f"chunk {batch_id[:20]}: {type(e).__name__}: {e}"
                    chunk_errors.append(msg)
                    print(f"{batch_name:<40} {provider:<12} {'chunk-error':<15} {msg}",
                          file=sys.stderr)
                    continue

                b_status = batch_resp.get("status", "unknown")
                counts = batch_resp.get("request_counts", {}) or {}
                cc = counts.get("completed", 0)
                tc = counts.get("total", 0)
                completed_count += cc
                total_count += tc

                # If a batch reached a non-completed terminal state, capture
                # OpenAI's failure detail so the operator can see what happened.
                # OpenAI returns errors in batch_resp["errors"]["data"] (a list).
                error_detail = None
                if b_status in ("failed", "cancelled", "expired"):
                    err_obj = batch_resp.get("errors")
                    if err_obj:
                        # First error message is enough to identify the cause
                        # (e.g., "insufficient_quota", "validation_error").
                        try:
                            first = err_obj.get("data", [{}])[0]
                            error_detail = (f"{first.get('code', 'unknown')}: "
                                            f"{first.get('message', '')[:200]}")
                        except Exception:
                            error_detail = str(err_obj)[:200]

                # Record per-chunk progress so subsequent polls do not lose
                # this information if the next poll hits a transient failure.
                chunk_progress[batch_id] = {
                    "status": b_status,
                    "output_file_id": batch_resp.get("output_file_id"),
                    "completed_requests": cc,
                    "total_requests": tc,
                    "error_detail": error_detail,
                }

            # Aggregate bundle status. Three possible terminal states:
            #   completed: every chunk completed with output_file_id
            #   completed_with_failures: some chunks completed, others terminal-failed
            #   failed: all chunks reached a non-completed terminal state
            # Non-terminal states (in_progress/finalizing/validating) keep the
            # bundle in "submitted" so the orchestrator continues polling.
            TERMINAL = {"completed", "failed", "cancelled", "expired"}
            active_ids = [bid for bid in info.get("batch_ids", [])
                          if not bid.startswith("ERROR")]
            chunk_states = [chunk_progress.get(bid, {}).get("status")
                             for bid in active_ids]
            all_terminal = active_ids and all(s in TERMINAL for s in chunk_states)
            n_completed = sum(1 for s in chunk_states if s == "completed")
            n_failed = sum(1 for s in chunk_states
                            if s in ("failed", "cancelled", "expired"))

            if all_terminal:
                if n_completed == len(active_ids):
                    info["status"] = "completed"
                elif n_completed > 0:
                    info["status"] = "completed_with_failures"
                    info["failed_chunks"] = n_failed
                else:
                    info["status"] = "failed"
                    info["failed_chunks"] = n_failed
                # Save the output_file_ids of any chunks that DID complete so
                # download can still recover their data, even if other chunks
                # of the same bundle failed.
                info["output_file_ids"] = [
                    chunk_progress[bid].get("output_file_id")
                    for bid in active_ids
                    if chunk_progress.get(bid, {}).get("status") == "completed"
                    and chunk_progress.get(bid, {}).get("output_file_id")
                ]

            details = f"{completed_count}/{total_count} requests"
            if n_failed:
                details += f" ({n_failed} chunk(s) terminal-failed)"
            if chunk_errors:
                details += f" ({len(chunk_errors)} poll error(s) — will retry)"
            status_label = info.get("status", "in_progress") if all_terminal else "in_progress"
            print(f"{batch_name:<40} {provider:<12} {status_label:<15} {details}")

            # Surface OpenAI failure reasons to stderr so the orchestrator log
            # captures them (run_pilot.py's run_pipeline forwards stderr).
            for bid in active_ids:
                cp = chunk_progress.get(bid, {})
                if cp.get("status") in ("failed", "cancelled", "expired") and cp.get("error_detail"):
                    print(f"  ✗ {bid[:24]}: {cp['error_detail']}", file=sys.stderr)

        # Check Anthropic batch status
        elif provider == "anthropic":
            try:
                api_key = config.get_api_key("anthropic")
                # Track per-batch state explicitly so the bundle is only
                # marked "ended" when EVERY chunk has reached "ended".
                # Otherwise, with multi-chunk submissions (added 2026-05-13
                # to handle the 256 MB body limit), an early-finishing chunk
                # would prematurely flip the bundle status and cause
                # cmd_download to attempt fetching unfinished batches.
                batch_progress = info.setdefault("batch_progress", {})
                batch_ids = info.get("batch_ids", [])
                ended_count = 0
                for batch_id in batch_ids:
                    # Skip re-polling batches we already know are ended.
                    prev = batch_progress.get(batch_id, {})
                    if prev.get("processing_status") == "ended":
                        ended_count += 1
                        continue

                    def _poll_anthropic_batch(_bid=batch_id):
                        req = urllib.request.Request(
                            f"https://api.anthropic.com/v1/messages/batches/{_bid}",
                            headers={
                                "x-api-key": api_key,
                                "anthropic-version": "2023-06-01",
                            },
                        )
                        with urllib.request.urlopen(req) as resp:
                            return json.loads(resp.read())

                    batch_resp = _openai_call_with_retries(
                        _poll_anthropic_batch,
                        _label=f"poll {batch_name}/{batch_id[:20]}")
                    b_status = batch_resp.get("processing_status", "unknown")
                    counts = batch_resp.get("request_counts", {}) or {}
                    batch_progress[batch_id] = {
                        "processing_status": b_status,
                        "succeeded": counts.get("succeeded", 0),
                        "errored": counts.get("errored", 0),
                        "canceled": counts.get("canceled", 0),
                        "expired": counts.get("expired", 0),
                        "processing": counts.get("processing", 0),
                    }
                    if b_status == "ended":
                        ended_count += 1
                    print(f"{batch_name:<40} {provider:<12} {b_status:<15} "
                          f"({batch_id[:24]}: {counts})")

                # Aggregate bundle status: only "ended" if ALL batches have ended.
                if batch_ids and ended_count == len(batch_ids):
                    info["status"] = "ended"
                    print(f"  → {batch_name}: all {len(batch_ids)} batches ended")
            except Exception as e:
                print(f"{batch_name:<40} {provider:<12} {'error':<15} {e}")

        elif provider == "google":
            print(f"{batch_name:<40} {provider:<12} {status:<15} sync (already done)")

    # Persist updated chunk_progress / status (atomic — no torn writes)
    atomic_write_json(tracking_path, tracking)


# ============================================================
# Download Command
# ============================================================

def cmd_download(args):
    """Download results from completed batches."""
    import urllib.request

    tracking_path = config.RESULTS_DIR / "batch_tracking.json"
    if not tracking_path.exists():
        print("No batches submitted yet.")
        return

    with open(tracking_path) as f:
        tracking = json.load(f)

    config.RESPONSES_DIR.mkdir(parents=True, exist_ok=True)

    for batch_name, info in tracking.items():
        provider = info["provider"]

        # Download whichever chunks completed, even if the bundle had partial
        # failures. info.output_file_ids was set by cmd_status to include only
        # successful chunks' output files, so this loop is safe.
        if provider == "openai" and info.get("status") in ("completed", "completed_with_failures"):
            for i, file_id in enumerate(info.get("output_file_ids", [])):
                out_path = config.RESPONSES_DIR / f"{batch_name}_chunk{i:03d}_results.jsonl"
                if out_path.exists():
                    print(f"{batch_name} chunk {i}: Already downloaded")
                    continue

                def _dl_openai_chunk():
                    api_key = config.get_api_key("openai")
                    req = urllib.request.Request(
                        f"https://api.openai.com/v1/files/{file_id}/content",
                        headers={"Authorization": f"Bearer {api_key}"},
                    )
                    with urllib.request.urlopen(req) as resp:
                        return resp.read()

                try:
                    data = _openai_call_with_retries(
                        _dl_openai_chunk,
                        _label=f"download {batch_name} chunk {i}")
                    # Sanity check: response should be non-empty JSONL
                    if not data or len(data) < 10:
                        raise RuntimeError(
                            f"download returned suspiciously small payload "
                            f"({len(data)} bytes) — refusing to write")
                    atomic_write_bytes(out_path, data)
                    print(f"{batch_name} chunk {i}: Downloaded to {out_path}")
                except Exception as e:
                    print(f"{batch_name} chunk {i}: ERROR - {e}", file=sys.stderr)

        elif provider == "anthropic" and info.get("status") == "ended":
            # Each batch_id gets a UNIQUE output filename so multi-chunk
            # submissions don't overwrite each other. The naming convention
            # mirrors OpenAI's chunked outputs so parse_response_filename
            # recognises them automatically.
            for i, batch_id in enumerate(info.get("batch_ids", [])):
                out_path = (config.RESPONSES_DIR /
                            f"{batch_name}_chunk{i:03d}_results.jsonl")
                if out_path.exists():
                    print(f"{batch_name} chunk {i}: Already downloaded")
                    continue

                def _dl_anthropic(_bid=batch_id):
                    api_key = config.get_api_key("anthropic")
                    req = urllib.request.Request(
                        f"https://api.anthropic.com/v1/messages/batches/{_bid}/results",
                        headers={
                            "x-api-key": api_key,
                            "anthropic-version": "2023-06-01",
                        },
                    )
                    with urllib.request.urlopen(req) as resp:
                        return resp.read()

                try:
                    data = _openai_call_with_retries(
                        _dl_anthropic,
                        _label=f"download {batch_name} chunk {i}")
                    if not data or len(data) < 10:
                        raise RuntimeError(
                            f"download returned suspiciously small payload "
                            f"({len(data)} bytes) — refusing to write")
                    atomic_write_bytes(out_path, data)
                    print(f"{batch_name} chunk {i}: Downloaded to {out_path}")
                except Exception as e:
                    print(f"{batch_name} chunk {i}: ERROR - {e}", file=sys.stderr)

        elif provider == "google":
            print(f"{batch_name}: Results already saved during submission")

        else:
            print(f"{batch_name}: Not ready (status: {info.get('status')})")


# ============================================================
# Parse Command
# ============================================================

def parse_grid_coordinate(text):
    """
    Parse grid coordinates from model response text.

    Handles: "B3", "B 3", "b3", "B-3", "Row B, Column 3", "(B3)", "B3.", etc.

    Returns list of normalized coordinates like ["B3"] or ["C4", "D5", "E6"],
    or None if unparseable.
    """
    import re

    if not text:
        return None

    text = text.strip()

    # Pattern: letter A-H followed by number 1-16
    pattern = r'[A-Ha-h]\s*[-]?\s*(?:1[0-6]|[1-9])\b'
    matches = re.findall(pattern, text)

    if not matches:
        # Try "Row X, Column Y" format
        alt_pattern = r'[Rr]ow\s*([A-Ha-h])\s*,?\s*[Cc]ol(?:umn)?\s*(\d{1,2})'
        alt_matches = re.findall(alt_pattern, text)
        if alt_matches:
            matches = [f"{r}{c}" for r, c in alt_matches]

    if not matches:
        return None

    # Normalize: uppercase, no spaces/dashes
    normalized = []
    for m in matches:
        clean = re.sub(r'[\s\-]', '', m).upper()
        normalized.append(clean)

    # Deduplicate while preserving order
    seen = set()
    unique = []
    for c in normalized:
        if c not in seen:
            seen.add(c)
            unique.append(c)

    return unique


_STRONG_REFUSAL_MARKERS = (
    # Require refusal + action object to avoid matching hedged phrases like
    # "I cannot be sure" (which is ambiguous, not refusal).
    "cannot identify", "can't identify",
    "cannot determine", "can't determine",
    "cannot provide", "can't provide",
    "cannot assist", "can't assist",
    "cannot help", "can't help",
    "cannot locate", "can't locate",
    "cannot answer", "can't answer",
    "cannot give", "can't give",
    "unable to identify", "unable to determine", "unable to provide",
    "unable to assist", "unable to help", "unable to locate", "unable to answer",
    "not able to identify", "not able to determine", "not able to provide",
    "as an ai", "as a language model",
    "i apologize", "i'm sorry", "sorry, i",
    "not qualified", "medical advice",
    "i refuse", "i must decline", "decline to",
)

_AMBIGUOUS_MARKERS = (
    "either", "or possibly", "might be", "could be", "possibly",
    "probably", "likely", "perhaps", "approximately",
    "not sure", "unclear", "hard to tell", "difficult to determine",
    "multiple possible", "several possible", "any of", "one of",
    "without more", "without additional", "cannot be sure", "can't be sure",
    "cannot be certain", "can't be certain", "it's hard", "unclear which",
)


def categorise_unparseable_response(text):
    """
    Classify a response that failed coordinate parsing (or was dropped as
    invalid) into one of four failure modes. Rule-based, deterministic,
    no LLM calls.

    Returns one of:
      - "no_engage" : empty / whitespace / near-empty response
      - "ambiguous" : hedged with multiple candidates or non-committal
                      language (checked BEFORE refusal to avoid the
                      "I cannot be sure" false-positive)
      - "refusal"   : explicit refusal or safety disclaimer
      - "verbose"   : substantive answer but no valid grid coordinates
    """
    if text is None:
        return "no_engage"
    stripped = text.strip()
    if len(stripped) < 5:
        return "no_engage"

    lower = stripped.lower()

    # Ambiguity is checked first: hedged refusal phrases like "I cannot
    # be sure" should land in "ambiguous", not "refusal".
    if any(m in lower for m in _AMBIGUOUS_MARKERS):
        return "ambiguous"
    if any(m in lower for m in _STRONG_REFUSAL_MARKERS):
        return "refusal"
    return "verbose"


def compute_compliance_stats(records):
    """
    Compute instruction compliance rate and failure-mode breakdown from a
    list of parsed response records.

    A record is considered compliant only if parsed_coordinates is non-empty
    AND failure_category is None (i.e., no out-of-range / ambiguous override).
    """
    total = len(records)
    parsed_ok = 0
    categories = {
        "refusal": 0, "verbose": 0, "ambiguous": 0,
        "no_engage": 0, "out_of_range": 0,
    }

    for r in records:
        if r.get("parsed_coordinates") and r.get("failure_category") is None:
            parsed_ok += 1
        else:
            cat = r.get("failure_category")
            if cat in categories:
                categories[cat] += 1

    compliance_rate = (parsed_ok / total) if total else 0.0
    return {
        "total": total,
        "compliant": parsed_ok,
        "non_compliant": total - parsed_ok,
        "compliance_rate": compliance_rate,
        "failure_modes": categories,
    }


def validate_coordinate(coord, modality):
    """
    Check if a coordinate is within the valid grid range for a modality.
    Returns True if valid, False if out of range or malformed.
    """
    import re
    if coord is None:
        return False
    coord = str(coord).strip().upper()
    match = re.match(r'^([A-H])(\d{1,2})$', coord)
    if not match:
        return False
    row_letter = match.group(1)
    col_num = int(match.group(2))
    grid = config.GRID_SPECS.get(modality)
    if not grid:
        return False
    max_row = grid["max_row_letter"]
    max_col = grid["cols"]
    # Explicit membership check — do not rely on ASCII lex order alone.
    valid_rows = "".join(chr(ord('A') + i) for i in range(ord(max_row) - ord('A') + 1))
    return row_letter in valid_rows and 1 <= col_num <= max_col


def parse_response_filename(name):
    """
    Identify (model_key, strategy) from a response filename.

    Expected patterns (all start with `{model_key}_{strategy}`):
      OpenAI    : {model_key}_{strategy}_chunk{N:03d}_results.jsonl
      Google    : {model_key}_{strategy}_chunk{N:03d}.json
      Anthropic : {model_key}_{strategy}_chunk{N:03d}_results.jsonl

    Returns (model_key, strategy) or (None, None) if no pattern matches.

    Uses config.ALL_STRATEGIES (the FULL strategy registry — canonical +
    ablation variants), not config.STRATEGIES (which is just the active
    submit-by-default pair). Without this, ablation filenames like
    `gpt-5.4_guided_no_tooth_num_chunk000_results.jsonl` would mis-attribute
    to the shorter "guided" prefix and the parsed `strategy` field on each
    record would silently collapse to "guided". Longest-first iteration
    ensures `guided_no_tooth_num` is tried before `guided`.
    """
    for mk in sorted(config.MODELS.keys(), key=len, reverse=True):
        for strat in sorted(config.ALL_STRATEGIES, key=len, reverse=True):
            prefix = f"{mk}_{strat}"
            if name == prefix or name.startswith(prefix + "_") or name.startswith(prefix + "."):
                return mk, strat
    return None, None


def _finalise_record(record, query_modality_map, query_landmark_map):
    """
    Validate parsed_coordinates in-place against the modality grid and enforce
    the single-cell rule for point landmarks. Mutates `record` and returns it.

    Rules (in order):
      1. If parsed_coordinates is None/empty, derive a failure_category from
         the raw response text (refusal / ambiguous / verbose / no_engage).
      2. If the modality is unknown, preserve the parse and flag nothing
         (this shouldn't happen in production — query_index always has the
         sheet name — but is preserved for defensive testing).
      3. POINT LANDMARK: the prompt asks for "exactly one cell coordinate".
         Any response that yields more than one cell — even if some happen
         to be out of grid range — is treated as "ambiguous". The model
         violated the instruction, full stop. This is the strict scientific
         interpretation: we don't second-guess which of the multiple cells
         was the model's "real" answer.
         A single coord that is out of range is "out_of_range".
      4. AREA LANDMARK: drop out-of-range cells silently and keep the valid
         subset. If no valid cells remain, classify as "out_of_range".

    Out-of-range cells (dropped or rejected) are preserved on the record in
    `record["out_of_range"]` for auditability.
    """
    query_id = record["query_id"]
    modality = query_modality_map.get(query_id)
    lm_type = query_landmark_map.get(query_id)
    record["modality"] = modality
    record["landmark_type"] = lm_type

    raw = record.get("raw_response")
    parsed = record.get("parsed_coordinates")

    # Branch 1: nothing parsed
    if not parsed:
        record["parsed_coordinates"] = None
        record["out_of_range"] = []
        record["failure_category"] = categorise_unparseable_response(raw)
        return record

    # Branch 2: modality unknown -> defensive passthrough
    if modality is None:
        record["out_of_range"] = []
        record["failure_category"] = None
        return record

    valid = [c for c in parsed if validate_coordinate(c, modality)]
    invalid = [c for c in parsed if not validate_coordinate(c, modality)]
    record["out_of_range"] = invalid

    # Branch 3: point landmark — strict single-cell rule
    if lm_type == "point":
        # Multi-coord response (any combination of valid+invalid) is
        # an instruction violation -> ambiguous.
        if len(parsed) > 1:
            record["parsed_coordinates"] = None
            record["failure_category"] = "ambiguous"
            return record
        # Exactly one parsed cell — succeeds iff it's in range.
        if not valid:
            record["parsed_coordinates"] = None
            record["failure_category"] = "out_of_range"
            return record
        record["parsed_coordinates"] = valid
        record["failure_category"] = None
        return record

    # Branch 4: area landmark — drop invalid cells, keep the valid subset
    if not valid:
        record["parsed_coordinates"] = None
        record["failure_category"] = "out_of_range"
        return record
    record["parsed_coordinates"] = valid
    record["failure_category"] = None
    return record


def cmd_parse(args):
    """
    Parse all downloaded responses into a single list of per-query records.

    Output file: results/parsed_responses.json

    Output format: a JSON list, one record per (model, strategy, query_id).
    Each record has explicit fields so downstream consumers never need to
    parse custom_ids or filenames:
        {
            "query_id": "PAN_001_Mental_Foramen_L",
            "strategy": "zero_shot",
            "model_key": "gpt-5.4",
            "modality": "PANORAMIC",
            "landmark_type": "point",
            "custom_id": "PAN_001_Mental_Foramen_L_zero_shot",
            "raw_response": "...",
            "parsed_coordinates": ["B3"],        # or None
            "out_of_range": [],                  # dropped cells
            "failure_category": null,            # or a string
            "source_file": "gpt-5.4_zero_shot_chunk000_results.jsonl",
        }

    This format fixes the prior custom_id-collision bug where responses from
    different providers overwrote each other in a dict keyed by custom_id.
    """
    # Build query_id -> modality and query_id -> landmark_type maps from the
    # query index. These are needed for validation and for M1 (single-cell
    # rule on point landmarks).
    index_path = config.RESULTS_DIR / "query_index.json"
    if not index_path.exists():
        print("ERROR: query_index.json not found. Run 'prepare' first.")
        sys.exit(1)
    with open(index_path) as f:
        queries = json.load(f)
    query_modality_map = {q["query_id"]: q["sheet"] for q in queries}
    query_landmark_map = {q["query_id"]: q["landmark_type"] for q in queries}

    records = []

    def add_record(model_key, strategy, custom_id, raw_text, source_name):
        # Strip off the strategy suffix from custom_id to recover query_id
        # (custom_id is built as "{query_id}_{strategy}" by the request builders).
        suffix = f"_{strategy}"
        query_id = custom_id[:-len(suffix)] if custom_id.endswith(suffix) else custom_id
        parsed = parse_grid_coordinate(raw_text)
        record = {
            "query_id": query_id,
            "strategy": strategy,
            "model_key": model_key,
            "custom_id": custom_id,
            "raw_response": raw_text,
            "parsed_coordinates": parsed,
            "source_file": source_name,
        }
        _finalise_record(record, query_modality_map, query_landmark_map)
        records.append(record)

    # ----- OpenAI / Anthropic JSONL files (sorted for determinism) -----
    for resp_file in sorted(config.RESPONSES_DIR.glob("*_results.jsonl")):
        model_key, strategy = parse_response_filename(resp_file.name)
        if model_key is None or strategy is None:
            print(f"WARNING: skipping unrecognised file {resp_file.name}")
            continue
        with open(resp_file) as f:
            for line in f:
                if not line.strip():
                    continue
                entry = json.loads(line)
                custom_id = entry.get("custom_id", "")

                response_text = None
                if "response" in entry and "body" in entry["response"]:
                    # OpenAI format
                    body = entry["response"]["body"]
                    choices = body.get("choices", [])
                    if choices:
                        response_text = choices[0].get("message", {}).get("content", "")
                elif "result" in entry:
                    # Anthropic format — guard against non-text content blocks.
                    result = entry["result"]
                    if result.get("type") == "succeeded":
                        content = result.get("message", {}).get("content", [])
                        if content and content[0].get("type") == "text":
                            response_text = content[0].get("text", "")

                add_record(model_key, strategy, custom_id, response_text, resp_file.name)

    # ----- Google JSON chunk files (sorted for determinism) -----
    for resp_file in sorted(config.RESPONSES_DIR.glob("*_chunk*.json")):
        model_key, strategy = parse_response_filename(resp_file.name)
        if model_key is None or strategy is None:
            print(f"WARNING: skipping unrecognised file {resp_file.name}")
            continue
        with open(resp_file) as f:
            entries = json.load(f)
        for entry in entries:
            custom_id = entry.get("custom_id", "")
            response = entry.get("response", {})
            response_text = None
            candidates = response.get("candidates", []) if response else []
            if candidates:
                parts = candidates[0].get("content", {}).get("parts", [])
                if parts:
                    response_text = parts[0].get("text", "")
            add_record(model_key, strategy, custom_id, response_text, resp_file.name)

    # Save unified results as a list of records (no more dict-keyed overwrites)
    # Atomic write: a crash mid-write cannot leave a torn parsed_responses.json
    # that downstream analysis would silently load.
    out_path = config.RESULTS_DIR / "parsed_responses.json"
    atomic_write_json(out_path, records)

    # Compute instruction compliance stats (Issues 5 + 8)
    compliance = compute_compliance_stats(records)
    compliance_path = config.RESULTS_DIR / "compliance_stats.json"
    atomic_write_json(compliance_path, compliance)

    # Per-(model, strategy) compliance breakdown
    from collections import defaultdict
    per_batch = defaultdict(list)
    for r in records:
        per_batch[(r["model_key"], r["strategy"])].append(r)

    total = compliance["total"]
    parsed_ok = compliance["compliant"]
    failed = compliance["non_compliant"]
    print(f"\nParsed {total} responses: {parsed_ok} OK, {failed} non-compliant")
    print(f"Instruction compliance rate: {compliance['compliance_rate']:.1%}")
    fm = compliance["failure_modes"]
    print(f"Failure modes: refusal={fm['refusal']}, verbose={fm['verbose']}, "
          f"ambiguous={fm['ambiguous']}, no_engage={fm['no_engage']}, "
          f"out_of_range={fm['out_of_range']}")

    if per_batch:
        print("\nPer-batch compliance:")
        for (mk, strat), rs in sorted(per_batch.items()):
            c = compute_compliance_stats(rs)
            print(f"  {mk} / {strat}: {c['compliant']}/{c['total']} "
                  f"({c['compliance_rate']:.1%})")

    if failed > 0:
        print("\nExample non-compliant entries (first 10):")
        shown = 0
        for r in records:
            if r.get("parsed_coordinates") and r.get("failure_category") is None:
                continue
            raw = (r.get("raw_response") or "")[:100]
            print(f"  [{r['failure_category']}] {r['custom_id']} "
                  f"({r['model_key']}): {raw!r}")
            shown += 1
            if shown >= 10:
                break

    print(f"\nResults saved: {out_path}")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Dental MLLM Benchmark Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Commands:
  prepare   Parse Excel, validate images, save query index (no API key needed)
  submit    Submit batches to APIs (requires API keys)
  status    Check batch processing status
  download  Download completed batch results
  parse     Parse responses into unified results file
        """,
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    p_prepare = subparsers.add_parser("prepare", help="Parse Excel, validate images, save query index")
    p_prepare.add_argument("--n-per-modality", type=int, default=None,
                           help="Keep only N deterministically-spaced images per modality (for pilot runs)")
    p_prepare.add_argument("--subset-images", type=str, default=None,
                           help="Explicit comma-separated image_ids to keep (overrides --n-per-modality)")
    p_prepare.add_argument("--models", type=str, default=None,
                           help="Comma-separated model_keys to include in the summary (informational)")

    p_prepare_v2 = subparsers.add_parser(
        "prepare_v2",
        help="Build extended query_index from Final Excel (consensus GT + 2 raters + students)")
    p_prepare_v2.add_argument("--excel", type=str, default=None,
                              help="Path to Final benchmark Excel "
                                   "(default: data/Final_Dental_MLLM_Benchmark_Data.xlsx)")
    p_prepare_v2.add_argument("--anchor-to", type=str, default=None,
                              help="Frozen GPT-5.4 run sandbox to cryptographically anchor "
                                   "raw JSONLs from (e.g. results_full)")

    p_prepare_abl = subparsers.add_parser(
        "prepare_ablation",
        help="Build a focused, filtered query_index for an ablation experiment")
    p_prepare_abl.add_argument("--excel", type=str, default=None,
                               help="Path to Final benchmark Excel")
    p_prepare_abl.add_argument("--structures", type=str, required=True,
                               help="Comma-separated structures to keep (e.g. Tooth_33_Apex)")
    p_prepare_abl.add_argument("--expected-count", type=int, default=None,
                               help="Refuse if filter doesn't yield this exact number")
    p_prepare_abl.add_argument("--anchor-to", type=str, default=None,
                               help="Frozen GPT-5.4 run to anchor relevant raw JSONLs from")
    p_prepare_abl.add_argument("--v2-index", type=str, default=None,
                               help="Path to v2 query_index.json (default: results_consensus/)")
    p_prepare_abl.add_argument("--label", type=str, default="fdi_ablation",
                               help="Label for the ablation experiment manifest")

    p_submit = subparsers.add_parser("submit", help="Submit batches to APIs")
    p_submit.add_argument("--models", type=str, default=None,
                          help="Comma-separated model_keys to submit (overrides config.ACTIVE_MODELS)")
    p_submit.add_argument("--strategies", type=str, default=None,
                          help="Comma-separated strategies to submit (overrides config.STRATEGIES). "
                               "Used by ablation experiments (e.g. --strategies guided_no_tooth_num).")

    subparsers.add_parser("status", help="Check batch processing status")
    subparsers.add_parser("download", help="Download completed batch results")
    subparsers.add_parser("parse", help="Parse responses into unified results file")

    args = parser.parse_args()

    commands = {
        "prepare": cmd_prepare,
        "prepare_v2": cmd_prepare_v2,
        "prepare_ablation": cmd_prepare_ablation,
        "submit": cmd_submit,
        "status": cmd_status,
        "download": cmd_download,
        "parse": cmd_parse,
    }
    commands[args.command](args)


def _load_dotenv_at_root() -> None:
    """Load .env from project root if present (no python-dotenv dependency)."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path) as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("export "):
                line = line[len("export "):]
            if "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip()
            if len(val) >= 2 and val[0] == val[-1] and val[0] in ("'", '"'):
                val = val[1:-1]
            if key and key not in os.environ:
                os.environ[key] = val


if __name__ == "__main__":
    _load_dotenv_at_root()
    main()
