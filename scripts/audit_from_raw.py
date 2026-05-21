"""
TRUE from-RAW-data verification: re-derive every headline number in v5 starting
from the Excel + raw JSONLs/chunks, with NO dependency on analysis.json /
phase_b.json / full_run_records.pkl (which are derived files).

This is the verification that proves v5's numbers are reproducible from the
absolute raw source. Any discrepancy between this script's output and v5's
claims is a real bug in the derivation chain.

Reads (raw-only):
    data/Final_Dental_MLLM_Benchmark_Data.xlsx                      — consensus_gt
    results_full/run{1,2,3}/responses/*.jsonl                       — GPT raw
    results_full_gemini/run{1,2,3}/responses/*.json                 — Gemini raw
    results_full_gemini/run1/responses/gemini-3.1-pro_requeries.jsonl — re-queries

Compares to v5 docx claims and to analysis JSON outputs (sanity check on the
full derivation chain).
"""
from __future__ import annotations
import json, math, re, sys
from pathlib import Path
from collections import defaultdict, Counter
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import pipeline  # only for the parser regex; we re-implement everything else


# ════════════════════════════════════════════════════════════════════
# STEP 1: Read RAW Excel → query_id × consensus_gt
# ════════════════════════════════════════════════════════════════════
print("=" * 70)
print("STEP 1: Reading Excel (data/Final_Dental_MLLM_Benchmark_Data.xlsx)")
print("=" * 70)

wb = openpyxl.load_workbook(ROOT / "data" / "Final_Dental_MLLM_Benchmark_Data.xlsx",
                              data_only=True)
gt_by_qid: dict[str, dict] = {}
for sheet_name in ("PANORAMIC", "PERIAPICAL", "CEPHALOMETRIC"):
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    cols = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        img_id = str(row[cols["Image_ID"]]).strip()
        target = str(row[cols["Target_Structure"]]).strip()
        gt = row[cols["CONSENSUS_Ground_Truth"]]
        qid = f"{img_id}_{target}"
        gt_by_qid[qid] = {
            "consensus_gt": gt,
            "modality": sheet_name,
            "image_id": img_id,
            "structure": target,
        }
print(f"  Loaded {len(gt_by_qid)} queries with consensus_gt from Excel")

# Infer landmark_type and grid dims
AREA_LANDMARKS = {"Mandibular_Canal_L", "Maxillary_Sinus_R", "External_Oblique_Ridge_R"}
GRID_SPECS = {
    "PANORAMIC":     {"cols": 16, "rows": 8, "max_row": "H", "diagonal": math.hypot(16, 8)},
    "PERIAPICAL":    {"cols": 8,  "rows": 6, "max_row": "F", "diagonal": math.hypot(8, 6)},
    "CEPHALOMETRIC": {"cols": 10, "rows": 8, "max_row": "H", "diagonal": math.hypot(10, 8)},
}
for q in gt_by_qid.values():
    q["landmark_type"] = "area" if q["structure"] in AREA_LANDMARKS else "point"


# ════════════════════════════════════════════════════════════════════
# STEP 2: Parser primitives (re-implemented here, not imported)
# ════════════════════════════════════════════════════════════════════
def parse_cells_re(text: str | None) -> list[str]:
    """Re-implement the cell parser from scratch. Returns list of normalized cells."""
    if not text:
        return []
    pattern = r'[A-Ha-h]\s*[-]?\s*(?:1[0-6]|[1-9])\b'
    matches = re.findall(pattern, text.strip())
    if not matches:
        alt_pattern = r'[Rr]ow\s*([A-Ha-h])\s*,?\s*[Cc]ol(?:umn)?\s*(\d{1,2})'
        alt_matches = re.findall(alt_pattern, text)
        if alt_matches:
            matches = [f"{r}{c}" for r, c in alt_matches]
    normalized = []
    seen = set()
    for m in matches:
        clean = re.sub(r'[\s\-]', '', m).upper()
        if clean not in seen:
            seen.add(clean)
            normalized.append(clean)
    return normalized


def cell_to_xy(cell: str, modality: str) -> tuple[int, int] | None:
    """'C5' → (col=5, row=3) 1-indexed, validated against grid."""
    s = (cell or "").strip().upper()
    m = re.match(r'^([A-H])(\d{1,2})$', s)
    if not m:
        return None
    row_letter, col_str = m.group(1), m.group(2)
    col = int(col_str)
    row = ord(row_letter) - ord('A') + 1
    grid = GRID_SPECS.get(modality)
    if not grid:
        return None
    max_col, max_row_letter = grid["cols"], grid["max_row"]
    max_row = ord(max_row_letter) - ord('A') + 1
    if 1 <= col <= max_col and 1 <= row <= max_row:
        return (col, row)
    return None


def euclidean(p: tuple[int, int], q: tuple[int, int]) -> float:
    return math.hypot(p[0] - q[0], p[1] - q[1])


def jaccard(a: set, b: set) -> float:
    if not a and not b: return 1.0
    if not a or not b:  return 0.0
    return len(a & b) / len(a | b)


def metric_for_query(pred_cells_raw: list[str], gt_str: str | None,
                     modality: str, lm_type: str) -> dict:
    """Return {'ed': float|None, 'jaccard': float|None, 'failure': str|None}."""
    if not pred_cells_raw:
        # No cells parsed at all
        return {"ed": None, "jaccard": None, "failure": "no_parse"}

    # Single-cell rule for points
    if lm_type == "point" and len(pred_cells_raw) > 1:
        return {"ed": None, "jaccard": None, "failure": "ambiguous_multi"}

    # Validate cells
    pred_xy = [cell_to_xy(c, modality) for c in pred_cells_raw]
    valid_xy = [x for x in pred_xy if x is not None]
    if not valid_xy:
        return {"ed": None, "jaccard": None, "failure": "out_of_range"}

    # Parse GT
    if not gt_str:
        return {"ed": None, "jaccard": None, "failure": "no_gt"}
    gt_cells_raw = parse_cells_re(gt_str)
    gt_xy = [cell_to_xy(c, modality) for c in gt_cells_raw]
    gt_valid = [x for x in gt_xy if x is not None]
    if not gt_valid:
        return {"ed": None, "jaccard": None, "failure": "no_gt_valid"}

    if lm_type == "point":
        return {"ed": euclidean(valid_xy[0], gt_valid[0]),
                "jaccard": None, "failure": None}
    else:
        return {"ed": None,
                "jaccard": jaccard(set(valid_xy), set(gt_valid)),
                "failure": None}


# ════════════════════════════════════════════════════════════════════
# STEP 3: Read RAW GPT JSONLs (results_full/run*/responses/*.jsonl)
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("STEP 3: Reading RAW GPT JSONLs (results_full/run{1,2,3}/responses/)")
print("=" * 70)

gpt_raw: dict[tuple[str, int], str | None] = {}  # (custom_id, rep_idx) → raw_text
gpt_dir = ROOT / "results_full"
n_lines = 0
for rep in (1, 2, 3):
    rep_dir = gpt_dir / f"run{rep}" / "responses"
    for jl_path in sorted(rep_dir.glob("*.jsonl")):
        for line in open(jl_path):
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            cid = obj["custom_id"]
            # OpenAI format
            content = (obj.get("response", {}).get("body", {})
                       .get("choices", [{}])[0].get("message", {})
                       .get("content", "") or "").strip()
            gpt_raw[(cid, rep - 1)] = content
            n_lines += 1
print(f"  Loaded {n_lines} raw GPT response lines across 3 reps")
print(f"  Unique (custom_id, rep) keys: {len(gpt_raw)}")


# ════════════════════════════════════════════════════════════════════
# STEP 4: Read RAW Gemini chunks + requeries
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("STEP 4: Reading RAW Gemini chunks (results_full_gemini/run*/responses/)")
print("=" * 70)

gem_raw: dict[tuple[str, int], str | None] = {}
gem_dir = ROOT / "results_full_gemini"
n_gem_chunks = 0
for rep in (1, 2, 3):
    rep_dir = gem_dir / f"run{rep}" / "responses"
    for chunk_path in sorted(rep_dir.glob("gemini-3.1-pro_*_chunk*.json")):
        data = json.loads(chunk_path.read_text())
        for entry in data:
            cid = entry.get("custom_id", "")
            resp = entry.get("response") or {}
            cands = resp.get("candidates", []) or []
            text: str | None = None
            if cands and isinstance(cands[0], dict):
                parts = (cands[0].get("content", {}) or {}).get("parts", []) or []
                if parts and isinstance(parts[0], dict):
                    text = parts[0].get("text", "") or ""
            gem_raw[(cid, rep - 1)] = text
            n_gem_chunks += 1
    # Merge re-queries for rep 1
    rq_path = rep_dir / "gemini-3.1-pro_requeries.jsonl"
    if rq_path.exists():
        n_rq = 0
        for line in open(rq_path):
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            cid = obj["custom_id"]
            # Re-queries OVERWRITE chunk values for matching cids
            gem_raw[(cid, rep - 1)] = obj.get("raw_response", "")
            n_rq += 1
        print(f"  Rep {rep}: merged {n_rq} re-queries")
print(f"  Total raw Gemini response entries (post-merge): {len(gem_raw)}")


# ════════════════════════════════════════════════════════════════════
# STEP 5: Compute mean ED / mean Jaccard FROM RAW DATA per model
# ════════════════════════════════════════════════════════════════════
def compute_metrics_from_raw(raw_by_key: dict[tuple[str, int], str | None],
                              model_label: str) -> dict[str, float]:
    """Compute mean ED / mean Jaccard / compliance per modality × strategy × type."""
    out: dict[str, dict] = defaultdict(lambda: {"eds": [], "jaccs": [], "failures": 0})
    # Iterate over (qid, strategy) pairs
    for qid_info in gt_by_qid.values():
        for strat in ("zero_shot", "guided"):
            cid = f"{qid_info['image_id']}_{qid_info['structure']}_{strat}"
            mod = qid_info["modality"]
            lm_type = qid_info["landmark_type"]
            group = f"{mod}_{strat}_{lm_type}"
            # Aggregate over 3 reps → mean
            rep_metrics = []
            for rep_idx in range(3):
                raw = raw_by_key.get((cid, rep_idx))
                pred_cells = parse_cells_re(raw)
                m = metric_for_query(pred_cells, qid_info["consensus_gt"],
                                       mod, lm_type)
                if lm_type == "point":
                    rep_metrics.append(m["ed"])
                else:
                    rep_metrics.append(m["jaccard"])
            valid = [v for v in rep_metrics if v is not None]
            failed = 3 - len(valid)
            out[group]["failures"] += failed
            if valid:
                mean_v = sum(valid) / len(valid)
                if lm_type == "point":
                    out[group]["eds"].append(mean_v)
                else:
                    out[group]["jaccs"].append(mean_v)
    # Summarise per group
    summary = {}
    for g, d in out.items():
        if d["eds"]:
            summary[g] = {"mean_ed": sum(d["eds"]) / len(d["eds"]),
                           "n": len(d["eds"]), "failures_per_rep": d["failures"]}
        elif d["jaccs"]:
            summary[g] = {"mean_jaccard": sum(d["jaccs"]) / len(d["jaccs"]),
                           "n": len(d["jaccs"]), "failures_per_rep": d["failures"]}
    return summary


print()
print("=" * 70)
print("STEP 5: Computing mean ED / Jaccard FROM RAW for both models")
print("=" * 70)

gpt_from_raw = compute_metrics_from_raw(gpt_raw, "GPT-5.4")
gem_from_raw = compute_metrics_from_raw(gem_raw, "Gemini 3.1 Pro")

print(f"\n  {'Group':<35} {'GPT (raw)':<15} {'Gemini (raw)':<15} {'GPT-Gem':<10}")
for g in sorted(gpt_from_raw):
    metric = "mean_ed" if "point" in g else "mean_jaccard"
    gpt_v = gpt_from_raw[g].get(metric)
    gem_v = gem_from_raw[g].get(metric)
    diff = gpt_v - gem_v if (gpt_v is not None and gem_v is not None) else None
    print(f"  {g:<35} {gpt_v:>9.4f}        {gem_v:>9.4f}        {diff:>+9.4f}")


# ════════════════════════════════════════════════════════════════════
# STEP 6: Compare to analysis.json (sanity check derivation chain)
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("STEP 6: Compare RAW recompute to analysis.json (full derivation chain)")
print("=" * 70)

gpt_analysis = json.load(open(ROOT / "results_consensus" / "analysis.json"))
gem_analysis = json.load(open(ROOT / "results_full_gemini" / "analysis.json"))

n_match = 0
n_diff = 0
for g in sorted(gpt_from_raw):
    metric = "mean_ed" if "point" in g else "mean_jaccard"
    gpt_raw_v = gpt_from_raw[g][metric]
    gem_raw_v = gem_from_raw[g][metric]
    gpt_json_v = gpt_analysis["RQ1_modality_strategy"][g]["mean"]
    gem_json_v = gem_analysis["RQ1_modality_strategy"][g]["mean"]
    gpt_match = abs(gpt_raw_v - gpt_json_v) < 1e-6
    gem_match = abs(gem_raw_v - gem_json_v) < 1e-6
    flag_g = "✓" if gpt_match else "✗"
    flag_e = "✓" if gem_match else "✗"
    print(f"  [{flag_g}{flag_e}] {g}: GPT raw={gpt_raw_v:.4f} json={gpt_json_v:.4f}  "
          f"|  Gem raw={gem_raw_v:.4f} json={gem_json_v:.4f}")
    if gpt_match: n_match += 1
    else: n_diff += 1
    if gem_match: n_match += 1
    else: n_diff += 1
print(f"\n  Total comparisons: {n_match + n_diff}, matches: {n_match}, diffs: {n_diff}")


# ════════════════════════════════════════════════════════════════════
# STEP 7: F5/F6 attractor from RAW data
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("STEP 7: F5/F6 attractor for Tooth_33_Apex (PAN, guided) FROM RAW")
print("=" * 70)

for model_label, raw_by_key in [("GPT-5.4", gpt_raw), ("Gemini 3.1 Pro", gem_raw)]:
    for strat in ("zero_shot", "guided"):
        n_total = 0
        n_f56 = 0
        n_exact = 0
        for qid, info in gt_by_qid.items():
            if info["structure"] != "Tooth_33_Apex" or info["modality"] != "PANORAMIC":
                continue
            cid = f"{qid}_{strat}"
            gt = info["consensus_gt"]
            for rep_idx in range(3):
                raw = raw_by_key.get((cid, rep_idx))
                cells = parse_cells_re(raw)
                if len(cells) != 1:
                    continue  # ambiguous or empty
                cell = cells[0]
                xy = cell_to_xy(cell, "PANORAMIC")
                if xy is None:
                    continue
                n_total += 1
                if cell in ("F5", "F6"):
                    n_f56 += 1
                if cell == (gt or "").strip().upper():
                    n_exact += 1
        if n_total > 0:
            print(f"  {model_label}/Tooth_33_Apex/{strat}: "
                  f"{n_f56}/{n_total} ({n_f56/n_total*100:.1f}%) on F5/F6, "
                  f"{n_exact}/{n_total} ({n_exact/n_total*100:.1f}%) exact match")


# ════════════════════════════════════════════════════════════════════
# STEP 8: Compliance count from RAW
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("STEP 8: Compliance count FROM RAW")
print("=" * 70)

def compliance_from_raw(raw_by_key: dict[tuple[str, int], str | None], label: str):
    total = 0
    compliant = 0
    for qid_info in gt_by_qid.values():
        for strat in ("zero_shot", "guided"):
            cid = f"{qid_info['image_id']}_{qid_info['structure']}_{strat}"
            mod = qid_info["modality"]
            lm_type = qid_info["landmark_type"]
            for rep_idx in range(3):
                total += 1
                raw = raw_by_key.get((cid, rep_idx))
                if raw is None:
                    continue
                pred_cells = parse_cells_re(raw)
                m = metric_for_query(pred_cells, qid_info["consensus_gt"], mod, lm_type)
                if m["failure"] is None:
                    compliant += 1
    print(f"  {label}: {compliant}/{total} compliant ({compliant/total*100:.4f}%), "
          f"{total - compliant} failures")
    return total, compliant

compliance_from_raw(gpt_raw, "GPT-5.4")
compliance_from_raw(gem_raw, "Gemini 3.1 Pro")
